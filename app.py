import os
import gc
import json
import csv
import io
import hmac
import unicodedata
from datetime import datetime

from flask import Flask, jsonify, request, render_template, send_file, session, redirect, url_for
from flask_compress import Compress
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from werkzeug.utils import secure_filename
from sqlalchemy import func
import openpyxl

from models import db, RouteMonthly, AirportMonthly, Airport, Aircraft, UploadLog, ManualRoute, \
    AirlineMonthly, AirlineUploadLog, AirlineLoadFactorSnapshot, FuelSale, FuelSaleUploadLog, \
    AirportAlias, ProyeccionRuta, ProyeccionConfig, ProyeccionExclusion, MercadoMensual, \
    TipoCambioMensual, IndiceEconomico, AdminFile
from parser import parse_workbook, rows_to_monthly_records, MONTH_ORDER
from parser_aerolineas import extract_report
from geocode import COORDS, ARGENTINA_NAMES
import avion_model

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'dev-secret-change-me')
if app.secret_key == 'dev-secret-change-me' and os.environ.get('RENDER'):
    # Si SECRET_KEY no está seteada como env var en Render, las cookies de sesión quedan
    # firmadas con una clave pública y conocida (la de este mismo repo en GitHub), lo que
    # permite falsificar sesiones de admin. No debería llegarse acá nunca en producción.
    app.logger.warning("SECRET_KEY no está configurada como variable de entorno en Render. "
                        "Configurala en el dashboard antes de usar esto en producción.")

app.config.update(
    SESSION_COOKIE_SECURE=True,      # la cookie de sesión solo viaja por HTTPS
    SESSION_COOKIE_HTTPONLY=True,    # JS del navegador no puede leer la cookie (mitiga XSS)
    SESSION_COOKIE_SAMESITE='Lax',   # mitiga CSRF básico en requests cross-site
    MAX_CONTENT_LENGTH=25 * 1024 * 1024,  # 25MB por request: evita que un upload gigante
                                            # tumbe el worker por memoria en el free tier de Render
)
Compress(app)

# Render pone la app detrás de un proxy: sin esto, request.remote_addr sería siempre
# la IP interna del proxy y el rate limiting terminaría agrupando a todos los usuarios
# bajo una sola IP (o los dejaría a todos sin límite real).
from werkzeug.middleware.proxy_fix import ProxyFix
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1)

limiter = Limiter(
    app=app,
    key_func=get_remote_address,
    storage_uri="memory://",   # OK porque render.yaml corre 1 solo worker; si algún día
                                # se escala a >1 worker, esto necesita moverse a Redis.
    default_limits=["200 per hour"],
)


@app.errorhandler(429)
def _rate_limited(e):
    return jsonify({"error": "Demasiados intentos. Esperá un momento y volvé a intentar."}), 429

db_url = os.environ.get('DATABASE_URL', 'sqlite:///local_dev.db')
if db_url.startswith('postgres://'):
    db_url = db_url.replace('postgres://', 'postgresql://', 1)
app.config['SQLALCHEMY_DATABASE_URI'] = db_url
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {'pool_pre_ping': True}

db.init_app(app)

ADMIN_PASSWORD = os.environ.get('ADMIN_PASSWORD', 'changeme')
MAP_PASSWORD = os.environ.get('MAP_PASSWORD', 'changeme-map')
FUEL_PASSWORD = os.environ.get('FUEL_PASSWORD', 'changeme-fuel')  # segunda capa, solo para datos de combustible


def _passwords_match(supplied, expected):
    """Comparación segura contra timing attacks (en vez de == directo). Ambos deben
    ser str; se descarta si alguno falta para no comparar con None."""
    if not supplied or not expected:
        return False
    return hmac.compare_digest(supplied, expected)



_HISTORICAL_CACHE = None  # lazy in-memory cache: (route_rows, airport_rows), never toca la base
_HISTORICAL_DESCARTADAS = []  # filas del histórico descartadas por pax/vuelo imposible

_AIRLINE_SEATS_CACHE = None  # lazy in-memory cache de aerolineas_asientos.json


def get_airline_seats():
    """Dict {avion_generico: {aerolinea: {asientos, nota}}} cargado de aerolineas_asientos.json.
    Referencia estatica (no editable via admin) para overridear asientos_default segun aerolinea
    en el panel de detalle de ruta del mapa."""
    global _AIRLINE_SEATS_CACHE
    if _AIRLINE_SEATS_CACHE is None:
        path = os.path.join(os.path.dirname(__file__), 'aerolineas_asientos.json')
        if os.path.exists(path):
            with open(path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            data.pop('_meta', None)
            _AIRLINE_SEATS_CACHE = data
        else:
            _AIRLINE_SEATS_CACHE = {}
    return _AIRLINE_SEATS_CACHE

# Catálogo inicial de aviones, sembrado la primera vez que arranca la app.
# Valores tomados de los que antes estaban hardcodeados en avion_model.py.
DEFAULT_AIRCRAFT = [
    # name, tipo_fuselaje, consumo_hora_kg, velocidad_crucero_kmh, asientos_default
    ('Boeing 747',        'wide',   9300, 913, 400),
    ('Boeing 777',        'wide',   7200, 892, 317),
    ('Boeing 777-300ER',  'wide',   7200, 892, 317),
    ('Boeing 787',        'wide',   5150, 903, 300),
    ('Airbus A350',       'wide',   5800, 903, 325),
    ('Airbus A350-900',   'wide',   5800, 903, 325),
    ('Airbus A330',       'wide',   5800, 871, 269),
    ('Airbus A320',       'narrow', 2550, 830, 170),
    ('Airbus A321',       'narrow', 2700, 830, 170),
    ('Boeing 737',        'narrow', 2600, 830, 170),
]


@app.cli.command('init-db')
def init_db():
    with app.app_context():
        db.create_all()
        for name, (lat, lon) in COORDS.items():
            if not Airport.query.get(name):
                db.session.add(Airport(name=name, lat=lat, lon=lon, is_argentina=(name in ARGENTINA_NAMES)))
        db.session.commit()
    print("DB inicializada y aeropuertos sembrados.")


def ensure_tables():
    db.create_all()
    _ensure_manual_route_columns()
    _ensure_airport_columns()  # debe correr ANTES de cualquier Airport.query.*, si no esa
                                # misma consulta ya falla porque el modelo espera is_argentina
    _ensure_aircraft_columns()
    if Airport.query.count() == 0:
        for name, (lat, lon) in COORDS.items():
            db.session.add(Airport(name=name, lat=lat, lon=lon, is_argentina=(name in ARGENTINA_NAMES)))
        db.session.commit()
    else:
        _sync_airports_from_coords()
    if Aircraft.query.count() == 0:
        for name, tipo_fuselaje, consumo, velocidad, asientos in DEFAULT_AIRCRAFT:
            db.session.add(Aircraft(name=name, tipo_fuselaje=tipo_fuselaje,
                                     consumo_hora_kg=consumo, velocidad_crucero_kmh=velocidad,
                                     asientos_default=asientos))
        db.session.commit()
    _sync_aircraft_from_flota()


def _sync_airports_from_coords():
    """Inserta en la tabla Airport los lugares que estén en geocode.COORDS pero todavía no
    en la base. El seeding original solo corre con la tabla vacía, así que sin esto un
    aeropuerto agregado a geocode.py nunca llegaría a una base ya desplegada. Solo agrega:
    nunca pisa ni borra lo que el admin haya editado a mano."""
    existentes = {name for (name,) in db.session.query(Airport.name).all()}
    nuevos = [n for n in COORDS if n not in existentes]
    if not nuevos:
        return
    for name in nuevos:
        lat, lon = COORDS[name]
        db.session.add(Airport(name=name, lat=lat, lon=lon,
                               is_argentina=(name in ARGENTINA_NAMES)))
    db.session.commit()
    app.logger.info("Aeropuertos nuevos sembrados desde geocode.py: %s", nuevos)


def _sync_aircraft_from_flota():
    """Agrega a la tabla Aircraft los tipos de flota.json que todavía no estén ahí, con
    nombre EXACTO al de flota.json.

    El catálogo de este panel arrancó sembrado con los 10 aviones genéricos de
    DEFAULT_AIRCRAFT (de cuando el modelo de consumo era por hora). Desde que existe
    flota.json con 23 tipos calibrados contra la planilla real, ese catálogo se quedó
    incompleto — por eso solo mostraba 10 filas en vez de 23.

    Solo AGREGA lo que falte, nunca pisa una fila existente: si el admin ya editó a mano
    algún valor (por ejemplo el consumo/hora de un 747), esa edición no se toca.

    Para los tipos que coinciden por nombre con flota.json, el cálculo real de consumo NO
    sale de esta tabla — sale directo de flota.json (register_manual_route revisa flota
    primero). Los campos que se guardan acá (consumo_hora_kg, velocidad_crucero_kmh) son
    una aproximación derivada de a_t/b_kg_km, solo para que la fila no se vea vacía y sirva
    de referencia; no participan del cálculo cuando el nombre coincide con la flota."""
    existentes = {a.name for a in Aircraft.query.with_entities(Aircraft.name).all()}
    agregados = []
    for codigo, ficha in avion_model.get_flota().items():
        nombre = ficha.get('nombre')
        if not nombre or nombre in existentes or not ficha.get('en_escalera'):
            continue  # SR22 (piston, en_escalera=False) no es candidato de ruta comercial
        velocidad = avion_model._VEL_CRUCERO.get(ficha.get('fuselaje'), 830.0)
        b = ficha.get('b_kg_km') or 0.0
        # Aproximación: b (kg/km) x velocidad (km/h) ~= consumo en crucero estabilizado,
        # kg/h. No es exacto (b ya integra rodaje+ascenso+descenso repartidos en la
        # distancia) pero alcanza para que el campo muestre un valor con sentido físico.
        consumo_hora_aprox = round(b * velocidad, 1) if b else None
        tipo_fuselaje = ficha.get('fuselaje') or 'narrow'
        if tipo_fuselaje not in ('regional', 'narrow', 'wide', 'piston'):
            tipo_fuselaje = 'narrow'
        db.session.add(Aircraft(
            name=nombre, tipo_fuselaje=tipo_fuselaje,
            consumo_hora_kg=consumo_hora_aprox or 2600.0,
            velocidad_crucero_kmh=velocidad,
            asientos_default=ficha.get('asientos'),
        ))
        agregados.append(nombre)
    if agregados:
        db.session.commit()
        app.logger.info("Catálogo de aviones: agregados desde flota.json: %s", agregados)


def _ensure_manual_route_columns():
    """db.create_all() no altera tablas ya existentes — si manual_route ya existía
    (deploys previos) sin la columna consumo_kg_manual, la agrega acá. Idempotente
    y seguro en Postgres y SQLite; no requiere acceso a Shell en Render."""
    from sqlalchemy import inspect, text
    inspector = inspect(db.engine)
    if 'manual_route' not in inspector.get_table_names():
        return
    existing_cols = {c['name'] for c in inspector.get_columns('manual_route')}
    if 'consumo_kg_manual' not in existing_cols:
        dialect = db.engine.dialect.name
        col_type = 'DOUBLE PRECISION' if dialect == 'postgresql' else 'FLOAT'
        db.session.execute(text(f'ALTER TABLE manual_route ADD COLUMN consumo_kg_manual {col_type}'))
        db.session.commit()


def _ensure_airport_columns():
    """Si la tabla airport ya existía (deploys previos) sin la columna is_argentina, la agrega
    acá y backfillea True para los nombres que geocode.py marca como domésticos de Argentina."""
    from sqlalchemy import inspect, text
    inspector = inspect(db.engine)
    if 'airport' not in inspector.get_table_names():
        return
    existing_cols = {c['name'] for c in inspector.get_columns('airport')}
    if 'is_argentina' not in existing_cols:
        dialect = db.engine.dialect.name
        default_sql = 'FALSE' if dialect == 'postgresql' else '0'
        db.session.execute(text(
            f'ALTER TABLE airport ADD COLUMN is_argentina BOOLEAN NOT NULL DEFAULT {default_sql}'
        ))
        db.session.commit()
        if ARGENTINA_NAMES:
            Airport.query.filter(Airport.name.in_(ARGENTINA_NAMES)).update(
                {'is_argentina': True}, synchronize_session=False
            )
            db.session.commit()


def _ensure_aircraft_columns():
    """Si la tabla aircraft ya existía (deploys previos) sin peso_operativo_kg /
    elasticidad_override, las agrega acá. Ambas quedan NULL por defecto (opcionales) — el
    frontend cae a la aproximación genérica por tipo_fuselaje cuando no están cargadas."""
    from sqlalchemy import inspect, text
    inspector = inspect(db.engine)
    if 'aircraft' not in inspector.get_table_names():
        return
    existing_cols = {c['name'] for c in inspector.get_columns('aircraft')}
    dialect = db.engine.dialect.name
    col_type = 'DOUBLE PRECISION' if dialect == 'postgresql' else 'FLOAT'
    if 'peso_operativo_kg' not in existing_cols:
        db.session.execute(text(f'ALTER TABLE aircraft ADD COLUMN peso_operativo_kg {col_type}'))
        db.session.commit()
    if 'elasticidad_override' not in existing_cols:
        db.session.execute(text(f'ALTER TABLE aircraft ADD COLUMN elasticidad_override {col_type}'))
        db.session.commit()


@app.errorhandler(413)
def _request_too_large(e):
    return jsonify({"error": "El archivo es demasiado grande (máximo 25MB)."}), 413


@app.before_request
def _startup():
    if not getattr(app, '_tables_ready', False):
        ensure_tables()
        app._tables_ready = True


def transform_parsed_to_rows(parsed, known_airports):
    """Pure transform: parsed workbook dict -> (route_rows, airport_rows, unknown_airports).
    No toca la base — segura para usar con datos que solo viven en memoria (histórico)."""
    route_tables = {
        'cabotaje': parsed['cabotaje_pax'],
        'internacional': parsed['intl_pax'],
    }
    vuelos_tables = {
        'cabotaje': parsed['cabotaje_vuelos'],
        'internacional': parsed['intl_vuelos'],
    }

    # Ningún avión que vuele rutas argentinas supera esto de forma sostenida (ni el 747-8I,
    # el más grande de la flota, llega a 410 asientos) — una celda con más pasajeros por
    # vuelo que esto es un dato corrompido en el origen, no una ruta con mucho tráfico.
    # Encontrado en un backtest: 153 filas de 2015-2018 con `vuelos=1` y miles de pasajeros
    # (ej. Aeroparque-Iguazú, Oct-2016: 85.459 pax en "1" vuelo). No se puede saber cuál de
    # los dos números es el corrompido, así que se descarta la fila entera en vez de
    # arrastrar un promedio físicamente imposible al resto del sistema (estacionalidad,
    # backtests, asignación de avión por ocupación).
    MAX_PAX_POR_VUELO_PLAUSIBLE = 500
    route_rows = []
    unknown_airports = set()
    descartadas_por_calidad = []

    for tipo in ('cabotaje', 'internacional'):
        pax_recs = rows_to_monthly_records(route_tables[tipo])
        vuelos_recs = rows_to_monthly_records(vuelos_tables[tipo])

        pax_lookup = {}
        for name, year, month, val in pax_recs:
            if ' - ' not in name:
                continue
            o, d = [x.strip() for x in name.split(' - ', 1)]
            pax_lookup[(o, d, year, month)] = val

        vuelos_lookup = {}
        for name, year, month, val in vuelos_recs:
            if ' - ' not in name:
                continue
            o, d = [x.strip() for x in name.split(' - ', 1)]
            vuelos_lookup[(o, d, year, month)] = val

        keys = set(pax_lookup.keys()) | set(vuelos_lookup.keys())
        for (o, d, year, month) in keys:
            if o not in known_airports:
                unknown_airports.add(o)
            if d not in known_airports:
                unknown_airports.add(d)

            pax = pax_lookup.get((o, d, year, month))
            vuelos = vuelos_lookup.get((o, d, year, month))
            pax_v = round(pax * 1000) if pax else None
            vuelos_v = round(vuelos) if vuelos else None
            if pax_v is None and vuelos_v is None:
                continue

            if pax_v and vuelos_v and vuelos_v > 0 and (pax_v / vuelos_v) > MAX_PAX_POR_VUELO_PLAUSIBLE:
                descartadas_por_calidad.append({
                    'tipo': tipo, 'origin': o, 'dest': d, 'year': year, 'month': month,
                    'pax': pax_v, 'vuelos': vuelos_v,
                })
                continue

            route_rows.append({
                'tipo': tipo, 'origin': o, 'dest': d, 'year': year, 'month': month,
                'pax': pax_v, 'vuelos': vuelos_v
            })

    airport_rows = []
    airport_recs = rows_to_monthly_records(parsed['airport_pax'])
    for name, year, month, val in airport_recs:
        if val is None:
            continue
        airport_rows.append({
            'airport': name, 'year': year, 'month': month,
            'pax_total': round(val * 1000)
        })

    return route_rows, airport_rows, unknown_airports, descartadas_por_calidad


def get_historical_rows():
    """Carga el histórico 2001-2022 una sola vez, en memoria — nunca se escribe en la base."""
    global _HISTORICAL_CACHE
    if _HISTORICAL_CACHE is not None:
        return _HISTORICAL_CACHE

    historical_path = os.path.join(os.path.dirname(__file__), 'historical_2001_2022.json')
    if not os.path.exists(historical_path):
        _HISTORICAL_CACHE = ([], [])
        return _HISTORICAL_CACHE

    with open(historical_path, 'r', encoding='utf-8') as f:
        parsed = json.load(f)

    known_airports = {a.name for a in Airport.query.with_entities(Airport.name).all()}
    route_rows, airport_rows, _unknown, descartadas = transform_parsed_to_rows(parsed, known_airports)
    if descartadas:
        app.logger.warning(
            "Histórico 2001-2022: %d fila(s) descartadas por pax/vuelo físicamente "
            "imposible (revisar /api/cobertura_anac para el detalle). Ejemplos: %s",
            len(descartadas),
            ', '.join(f"{d['origin']}-{d['dest']} {d['year']}-{d['month']}" for d in descartadas[:5]))
    global _HISTORICAL_DESCARTADAS
    _HISTORICAL_DESCARTADAS = descartadas
    del parsed
    gc.collect()

    _HISTORICAL_CACHE = (route_rows, airport_rows)
    return _HISTORICAL_CACHE


def _get_upsert_builder():
    dialect = db.engine.dialect.name
    if dialect == 'postgresql':
        from sqlalchemy.dialects.postgresql import insert as db_insert
    elif dialect == 'sqlite':
        from sqlalchemy.dialects.sqlite import insert as db_insert
    else:
        db_insert = None
    return db_insert


def bulk_upsert(model, rows, conflict_cols, update_cols, batch_size=2000):
    if not rows:
        return
    db_insert = _get_upsert_builder()
    table = model.__table__
    for i in range(0, len(rows), batch_size):
        batch = rows[i:i + batch_size]
        if db_insert is not None:
            stmt = db_insert(table).values(batch)
            update_dict = {c: getattr(stmt.excluded, c) for c in update_cols}
            stmt = stmt.on_conflict_do_update(index_elements=conflict_cols, set_=update_dict)
            db.session.execute(stmt)
        else:
            for row in batch:
                db.session.merge(model(**row))
    db.session.commit()


def ingest_parsed_workbook(parsed, known_airports):
    """Usado solo por /upload (el archivo actual/incremental) — esto SI escribe en la base."""
    route_rows, airport_rows, unknown_airports, descartadas = transform_parsed_to_rows(parsed, known_airports)

    bulk_upsert(
        RouteMonthly, route_rows,
        conflict_cols=['tipo', 'origin', 'dest', 'year', 'month'],
        update_cols=['pax', 'vuelos']
    )
    n_routes = len(route_rows)
    del route_rows

    bulk_upsert(
        AirportMonthly, airport_rows,
        conflict_cols=['airport', 'year', 'month'],
        update_cols=['pax_total']
    )
    n_airports = len(airport_rows)
    del airport_rows

    return n_routes, n_airports, unknown_airports, descartadas


def get_aircraft_catalog():
    """Dict {nombre_avion: {consumo_hora_kg, velocidad_crucero_kmh, tipo_fuselaje,
    asientos_default, peso_operativo_kg, elasticidad_override}} para pasarle a
    avion_model.register_manual_route como fuente de verdad en vez de adivinar por texto.
    peso_operativo_kg y elasticidad_override viajan como null si no se cargaron (opcionales,
    el frontend cae a una aproximación genérica cuando faltan)."""
    return {
        a.name: {
            'consumo_hora_kg': a.consumo_hora_kg,
            'velocidad_crucero_kmh': a.velocidad_crucero_kmh,
            'tipo_fuselaje': a.tipo_fuselaje,
            'asientos_default': a.asientos_default,
            'peso_operativo_kg': a.peso_operativo_kg,
            'elasticidad_override': a.elasticidad_override,
        }
        for a in Aircraft.query.all()
    }


def _upsert_single_manual_route(o, d, tipo_clean, avion, asientos_int, consumo_kg_val,
                                 airports, aircraft_catalog):
    """Guarda/actualiza una fila de ManualRoute y refresca el cache de avion_model.
    Asume que o, d, avion, tipo_clean ya están validados/normalizados por el caller."""
    existing = ManualRoute.query.filter_by(origin=o, dest=d).first()
    if existing:
        existing.avion = avion
        existing.asientos = asientos_int
        existing.tipo = tipo_clean
        existing.consumo_kg_manual = consumo_kg_val
    else:
        db.session.add(ManualRoute(
            origin=o, dest=d, tipo=tipo_clean,
            avion=avion, asientos=asientos_int, consumo_kg_manual=consumo_kg_val
        ))

    avion_model.register_manual_route(o, d, tipo_clean, avion, asientos_int, airports,
                                       consumo_kg_override=consumo_kg_val,
                                       aircraft_catalog=aircraft_catalog)


def ingest_manual_routes(file_obj):
    """Parsea un Excel de rutas manuales (columnas: Partida, Arribo, Tipo, Avión, Asientos,
    Consumo (kg) [opcional]). Guarda en manual_route y registra el avión/consumo en avion_model
    cache. Si la columna Consumo viene vacía, el consumo se calcula automáticamente con el modelo.
    Devuelve (n_ok, errores)."""
    import openpyxl
    wb = openpyxl.load_workbook(file_obj, data_only=True)
    ws = wb.active

    airports = {a.name: (a.lat, a.lon) for a in Airport.query.all()}
    aircraft_catalog = get_aircraft_catalog()

    n_ok = 0
    errors = []

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue
        # Formato: Partida | Arribo | Tipo | Avión | Asientos | Consumo (kg) [opcional]
        partida, arribo, tipo, avion, asientos, consumo_kg = (row + (None,) * 6)[:6]

        if not partida or not arribo:
            errors.append(f"Fila {i}: Partida o Arribo vacíos")
            continue
        if not avion:
            errors.append(f"Fila {i}: Avión vacío")
            continue

        o = str(partida).strip()
        d = str(arribo).strip()
        tipo_clean = str(tipo).strip().lower() if tipo else 'internacional'
        if 'cab' in tipo_clean:
            tipo_clean = 'cabotaje'
        else:
            tipo_clean = 'internacional'

        unknown = []
        if o not in airports:
            unknown.append(o)
        if d not in airports:
            unknown.append(d)
        if unknown:
            errors.append(f"Fila {i}: lugar(es) sin coordenadas — {', '.join(unknown)}. "
                          f"Agregarlo a geocode.py primero.")
            continue

        asientos_int = None
        if asientos:
            try:
                asientos_int = int(asientos)
            except (ValueError, TypeError):
                errors.append(f"Fila {i}: Asientos inválido (no es un número): {asientos}")
                continue

        consumo_kg_val = None
        if consumo_kg not in (None, ''):
            try:
                consumo_kg_val = float(consumo_kg)
            except (ValueError, TypeError):
                errors.append(f"Fila {i}: Consumo (kg) inválido (no es un número): {consumo_kg}")
                continue

        _upsert_single_manual_route(o, d, tipo_clean, str(avion), asientos_int, consumo_kg_val,
                                     airports, aircraft_catalog)
        n_ok += 1

    db.session.commit()
    return n_ok, errors


# ---------- Public map (protegido con clave de acceso) ----------

def _normalize_airport_name(name):
    """Mayúsculas, sin tildes/diacríticos, espacios recortados. Usado para poder
    matchear 'EZEIZA' con 'Ezeiza' o 'CORDOBA' con 'Córdoba' sin necesitar un alias
    explícito para cada variante de mayúsculas/tildes."""
    if not name:
        return ''
    nfkd = unicodedata.normalize('NFKD', name)
    sin_tildes = ''.join(c for c in nfkd if not unicodedata.combining(c))
    return sin_tildes.strip().upper()


def resolve_airport_name(raw_name, known_airports_norm, alias_map):
    """Devuelve (nombre_canonico_o_None, matcheo: 'exact'|'normalized'|'alias'|None).
    known_airports_norm: dict {nombre_normalizado: Airport.name real}
    alias_map: dict {alias_normalizado: Airport.name real} (tabla AirportAlias)"""
    norm = _normalize_airport_name(raw_name)
    if norm in known_airports_norm:
        return known_airports_norm[norm], 'normalized'
    if norm in alias_map:
        return alias_map[norm], 'alias'
    return None, None


# Reemplazar la función `parse_fuel_sales_excel` existente en app.py por esta versión.
# Todo lo demás del archivo queda igual.
#
# BUG encontrado (agosto 2026): filas con VUELO vacío en el Excel de despachos se
# descartaban por completo, incluido su VOLUMEN. En el Excel de julio-2026 esto se comía
# 2.718,667 m3 reales (44 filas: 28 de Aviación General, 9 de Aerolíneas Argentinas, 4 de
# Traders Acuerdo, 1 de Copa, 1 de JetSmart) — el "Despacho YPF real" que se ve en
# Proyecciones sale directo de sumar FuelSale.volumen_m3, así que esas filas faltantes
# hacían parecer que el despacho real había caído, cuando en realidad estaba completo en el
# Excel y se perdía al subirlo.

def parse_fuel_sales_excel(file_obj):
    """Lee la hoja "Resumen" (o la primera hoja si no existe con ese nombre) generada
    por la macro de Excel. Columnas esperadas A-J, en este orden: VUELO, CANT VUELOS,
    AEROPUERTO, VOLUMEN, LLAA, DESTINO, MES, AÑO, PRECIO USD/L, INGRESO.
    read_only=True para mantener bajo el uso de memoria (mismo criterio que parser.py)."""
    wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
    ws = wb['Resumen'] if 'Resumen' in wb.sheetnames else wb.active

    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if not header:
        raise ValueError("La hoja está vacía")

    registros = []
    for row in rows_iter:
        if not row or all(v in (None, '') for v in row):
            continue  # fila realmente vacía (separador o cola de la hoja), no un despacho real
        vals = (list(row) + [None] * 10)[:10]
        vuelo, cant_vuelos, aeropuerto, volumen, llaa, destino, mes, anio, precio, ingreso = vals
        if not aeropuerto:
            continue  # sin aeropuerto no hay a qué asignar el volumen, esta fila sí se descarta
        registros.append({
            # VUELO vacío ya no descarta la fila: son despachos reales (aviación general,
            # acuerdos de trader, algún comercial sin nro. cargado en la macro) que antes se
            # perdían en silencio. Se etiquetan 'S/N' para no romper el agrupamiento
            # (vuelo, llaa) de la tabla del mapa.
            'vuelo': str(vuelo).strip() if vuelo not in (None, '') else 'S/N',
            'cant_vuelos': int(cant_vuelos) if cant_vuelos is not None else 0,
            'aeropuerto': str(aeropuerto).strip(),
            'volumen_m3': float(volumen) if volumen is not None else 0.0,
            'llaa': str(llaa).strip() if llaa is not None else '',
            'destino': str(destino).strip() if destino not in (None, '') else None,
            'mes': int(mes) if mes is not None else None,
            'anio': int(anio) if anio is not None else None,
            'precio_usd_l': float(precio) if isinstance(precio, (int, float)) else None,
            'ingreso_usd': float(ingreso) if ingreso is not None else 0.0,
        })
    return registros


MODEL_VERSION = 'asignación por ocupación'

@app.route('/')
def index():
    if not session.get('map_access'):
        return render_template('map_login.html')
    # no-store en map.html: el archivo se edita seguido y un navegador que se quede con la
    # copia vieja hace parecer que el deploy no salió. Pesa poco y solo es el HTML: los datos
    # ya viajan aparte por /api/data.
    resp = send_file('map.html')
    resp.headers['Cache-Control'] = 'no-store, must-revalidate'
    resp.headers['Pragma'] = 'no-cache'
    return resp


@app.route('/login', methods=['POST'])
@limiter.limit("10 per minute")
def map_login():
    if _passwords_match(request.form.get('password'), MAP_PASSWORD):
        session['map_access'] = True
        return redirect(url_for('index'))
    return render_template('map_login.html', error='Clave incorrecta.')


@app.route('/logout')
def map_logout():
    session.pop('map_access', None)
    session.pop('fuel_access', None)
    return redirect(url_for('index'))


@app.route('/fuel_login', methods=['POST'])
@limiter.limit("10 per minute")
def fuel_login():
    """Segunda capa de contraseña, independiente de la del mapa (MAP_PASSWORD): gatea
    específicamente todo lo que sale del Excel de ventas de combustible (panel, hover del
    mapa, tooltip del desplegable de aeropuerto). No requiere estar logueado como admin."""
    if not session.get('map_access'):
        return jsonify({"error": "No autorizado"}), 401
    if _passwords_match(request.form.get('password'), FUEL_PASSWORD):
        session['fuel_access'] = True
        return jsonify({"ok": True})
    return jsonify({"error": "Contraseña incorrecta"}), 401


@app.route('/api/data')
@limiter.limit("30 per minute")
def api_data():
    if not session.get('map_access'):
        return jsonify({"error": "No autorizado"}), 401
    try:
        hist_routes, hist_airports = get_historical_rows()

        db_routes = [
            {'tipo': r.tipo, 'origin': r.origin, 'dest': r.dest, 'year': r.year, 'month': r.month,
             'vuelos': r.vuelos, 'pax': r.pax}
            for r in RouteMonthly.query.all()
        ]
        db_airports = [
            {'airport': a.airport, 'year': a.year, 'month': a.month, 'pax_total': a.pax_total}
            for a in AirportMonthly.query.all()
        ]

        # Rutas manuales: pre-registrar en avion_model cache (avión/asientos/consumo confirmados)
        # y, solo si NO tienen tráfico real de ANAC, agregarlas como filas fijas (vuelos=1)
        # para que igual se dibujen en el mapa aunque no haya pasajeros/vuelos reales
        _all_airports = Airport.query.all()
        airports = {a.name: (a.lat, a.lon) for a in _all_airports}
        airports_is_argentina = {a.name: a.is_argentina for a in _all_airports}

        # Pares (origin, dest) que ya tienen tráfico real, para no duplicar vuelos
        real_traffic_pairs = set()
        for r in hist_routes + db_routes:
            real_traffic_pairs.add((r['origin'], r['dest']))
            real_traffic_pairs.add((r['dest'], r['origin']))  # por si el sentido está invertido

        # Computar todos los años disponibles una sola vez
        all_years = set()
        for r in hist_routes + db_routes:
            all_years.add(r['year'])
        all_years = all_years or ['2024']

        manual_route_rows = []
        aircraft_catalog = get_aircraft_catalog()
        for mr in ManualRoute.query.all():
            avion_model.register_manual_route(
                mr.origin, mr.dest, mr.tipo, mr.avion, mr.asientos, airports,
                consumo_kg_override=mr.consumo_kg_manual, aircraft_catalog=aircraft_catalog
            )
            if (mr.origin, mr.dest) in real_traffic_pairs:
                continue  # ya tiene vuelos/pax reales — no inyectar una fila ficticia encima
            # Inyectar en todos los años y meses disponibles como línea fija,
            # para que aparezca sin importar qué período esté mirando el usuario
            for yr in all_years:
                for mo in MONTH_ORDER:
                    manual_route_rows.append({
                        'tipo': mr.tipo, 'origin': mr.origin, 'dest': mr.dest,
                        'year': yr, 'month': mo,
                        'vuelos': 1, 'pax': None,
                    })

        all_routes = hist_routes + db_routes + manual_route_rows
        all_airport_rows = hist_airports + db_airports

        # ---- Pasada 1: rango de ocupación (pax por vuelo) observado en cada ruta ----
        # El avión ya no se elige por distancia sino por ocupación, y la ocupación cambia
        # mes a mes. Se calcula acá el rango histórico de cada ruta para mandarle al
        # frontend SOLO los tipos de avión que podrían llegar a elegirse, en vez de la
        # flota entera por cada una de las ~260 rutas.
        #
        # Se usan percentiles 5/95 y no el mínimo y máximo absolutos: una sola ruta con un
        # mes de arranque de 8 pax/vuelo en 2003 arrastraría toda la escalera desde el
        # CRJ200 y llenaría el payload de opciones que no se van a usar nunca.
        occ_vals = {}
        for r in all_routes:
            if not (r['pax'] and r['vuelos']):
                continue
            ppv = r['pax'] / r['vuelos']
            if ppv > 0:
                occ_vals.setdefault((r['origin'], r['dest']), []).append(ppv)

        def _pct(vals, p):
            if not vals:
                return None
            i = min(len(vals) - 1, max(0, int(round(p * (len(vals) - 1)))))
            return vals[i]

        occ_range = {}
        for k, vals in occ_vals.items():
            vals.sort()
            occ_range[k] = (_pct(vals, 0.05), _pct(vals, 0.95))
        del occ_vals

        cab_meta, cab_idx = [], {}
        intl_meta, intl_idx = [], {}
        cab_data, intl_data = {}, {}

        def get_idx(meta_list, idx_map, o, d, tipo):
            key = (o, d)
            if key in idx_map:
                return idx_map[key]
            if o not in airports or d not in airports:
                return None
            olat, olon = airports[o]
            dlat, dlon = airports[d]
            pax_min, pax_max = occ_range.get(key, (None, None))
            avion_info = avion_model.get_aircraft_info(o, d, tipo, airports)
            opciones = avion_model.opciones_para_rango(o, d, airports, pax_min, pax_max)
            idx_map[key] = len(meta_list)
            entry = [o, d, round(olat, 3), round(olon, 3), round(dlat, 3), round(dlon, 3)]
            if avion_info:
                entry += [
                    avion_info['avion'], avion_info['fuente'], avion_info['asientos'],
                    avion_info['distancia_km'], avion_info['consumo_total_kg'],
                    avion_info['consumo_total_m3'],
                ]
            else:
                entry += [None, None, None, None, None, None]
            # índice 12: opciones de avión para elegir en vivo según la ocupación del mes
            # [codigo, nombre, asientos, pax_vuelo_max, consumo_kg_base, fuselaje, es_real]
            entry.append(opciones)
            meta_list.append(entry)
            return idx_map[key]

        for r in all_routes:
            ym = f"{r['year']}-{r['month']}"
            ppv = round(r['pax'] / r['vuelos'], 1) if (r['pax'] and r['vuelos']) else None
            if r['tipo'] == 'cabotaje':
                idx = get_idx(cab_meta, cab_idx, r['origin'], r['dest'], 'cabotaje')
                if idx is None:
                    continue
                cab_data.setdefault(ym, []).append([idx, r['vuelos'], r['pax'], ppv])
            else:
                idx = get_idx(intl_meta, intl_idx, r['origin'], r['dest'], 'internacional')
                if idx is None:
                    continue
                intl_data.setdefault(ym, []).append([idx, r['vuelos'], r['pax'], ppv])

        node_names = set()
        for m in cab_meta + intl_meta:
            node_names.add(m[0])
            node_names.add(m[1])

        # Segunda capa de contraseña: ahora SOLO protege el dato de ingreso/facturación (USD).
        # Vuelos, despachos y volumen del Excel de combustible (incluidos los aeropuertos que
        # solo existen en el mapa por tener datos de combustible, sin ninguna ruta real de
        # pasajeros) se calculan y envían siempre, sin necesidad de desbloquear.
        fuel_access = bool(session.get('fuel_access'))

        fuel_sale_airport_names = {row[0] for row in db.session.query(FuelSale.aeropuerto).distinct().all()}
        isolated_airports = sorted(
            name for name in fuel_sale_airport_names
            if name in airports and name not in node_names
        )
        node_names |= set(isolated_airports)

        # Todo aeropuerto argentino conocido tiene que mandarse SIEMPRE, haya tenido o no
        # una ruta real de ANAC o un despacho de combustible alguna vez. Sin esto, un
        # aeropuerto que el usuario cargó a mano en "Aeropuertos conocidos" (por ejemplo
        # para asociarle despachos futuros) nunca aparecía en el mapa hasta el primer mes
        # en que efectivamente tuviera vuelos o combustible — "no voló este mes" no es lo
        # mismo que "no existe".
        node_names |= {name for name, es_ar in airports_is_argentina.items() if es_ar}

        airport_has_data = set(am['airport'] for am in all_airport_rows)

        node_meta, node_idx = [], {}
        for name in sorted(node_names):
            if name not in airports:
                continue
            lat, lon = airports[name]
            node_idx[name] = len(node_meta)
            node_meta.append([name, round(lat, 3), round(lon, 3), name in airport_has_data,
                               airports_is_argentina.get(name, False)])

        node_data = {}
        for am in all_airport_rows:
            if am['airport'] not in node_idx:
                continue
            ym = f"{am['year']}-{am['month']}"
            node_data.setdefault(ym, []).append([node_idx[am['airport']], am['pax_total']])

        # Datos de combustible agregados por aeropuerto+período (para el hover del punto en el
        # mapa y para el tooltip del desplegable "Aeropuerto", sin tener que abrir el panel).
        # [nodeIdx, vuelos_distintos, despachos_totales, volumen_m3, ingreso_usd]
        # El ingreso_usd va en null si la sesión no desbloqueó la segunda contraseña; el resto
        # de los campos (vuelos, despachos, volumen) se envían siempre.
        fuel_node_data = {}
        fuel_agg = {}  # (aeropuerto, anio, mes) -> {vuelos_set, despachos, volumen, ingreso}
        for r in FuelSale.query.all():
            key = (r.aeropuerto, r.anio, r.mes)
            entry = fuel_agg.setdefault(key, {'vuelos_set': set(), 'despachos': 0, 'volumen': 0.0, 'ingreso': 0.0})
            entry['vuelos_set'].add((r.vuelo, r.llaa))
            entry['despachos'] += r.cant_vuelos or 0
            entry['volumen'] += r.volumen_m3 or 0.0
            entry['ingreso'] += r.ingreso_usd or 0.0

        for (aeropuerto, anio, mes), entry in fuel_agg.items():
            if aeropuerto not in node_idx or not mes or mes < 1 or mes > 12:
                continue
            ym = f"{anio}-{MONTH_ORDER[mes - 1]}"
            fuel_node_data.setdefault(ym, []).append([
                node_idx[aeropuerto], len(entry['vuelos_set']), entry['despachos'],
                round(entry['volumen'], 3), round(entry['ingreso'], 2) if fuel_access else None,
            ])

        # years/year_months para el selector de período: solo a partir de tráfico REAL de ANAC
        # (hist_routes + db_routes), sin las rutas manuales inyectadas — si no, un mes futuro
        # que solo tiene una ruta manual ficticia (vuelos=1) terminaría eligiéndose como
        # "el mes más reciente" en vez del último mes con datos reales.
        real_yms = set()
        for r in hist_routes + db_routes:
            real_yms.add(f"{r['year']}-{r['month']}")
        years = sorted(set(ym.split('-')[0] for ym in real_yms), key=lambda y: int(y))
        year_months = {}
        for ym in real_yms:
            y, m = ym.split('-')
            year_months.setdefault(y, set()).add(m)
        for y in year_months:
            year_months[y] = sorted(year_months[y], key=lambda m: MONTH_ORDER.index(m))

        return jsonify({
            "years": years,
            "month_order": MONTH_ORDER,
            "year_months": year_months,
            "cabotaje": {"meta": cab_meta, "data": cab_data},
            "internacional": {"meta": intl_meta, "data": intl_data},
            "nodes": {"meta": node_meta, "data": node_data},
            "fuel_nodes": fuel_node_data,
            "fuel_access": fuel_access,
            "isolated_airports": isolated_airports,
            "aircraft_catalog": get_aircraft_catalog(),
            "airline_seats": get_airline_seats(),
            "flota": avion_model.get_flota(),
            "version": MODEL_VERSION,
            # Diagnóstico de deploy: si alguno de estos da 0, los JSON del modelo no llegaron
            # al servidor y el mapa estaría cayendo a estimaciones sin calibrar.
            "modelo_datos": {
                "tipos_en_flota": len(avion_model.get_flota()),
                "rutas_con_consumo_real": len(avion_model.get_matriz()),
            },
            "modelo_params": {
                "lf_referencia": avion_model.LF_REFERENCIA,
                "elasticidad": avion_model.FUEL_WEIGHT_ELASTICITY,
                "pax_weight_share": avion_model.PAX_WEIGHT_SHARE,
                "kg_por_m3": avion_model.JETA1_KG_PER_M3,
            },
        })
    except Exception as e:
        import traceback
        error_msg = str(e)
        stack = traceback.format_exc()
        print(f"Error en /api/data: {error_msg}\n{stack}", flush=True)
        return jsonify({"error": error_msg, "type": type(e).__name__}), 500


# ---------- Admin ----------

@app.route('/admin', methods=['GET'])
def admin_page():
    if not session.get('is_admin'):
        return render_template('login.html')
    logs = UploadLog.query.order_by(UploadLog.uploaded_at.desc()).limit(20).all()
    airline_logs = AirlineUploadLog.query.order_by(AirlineUploadLog.uploaded_at.desc()).limit(20).all()
    fuel_sale_logs = FuelSaleUploadLog.query.order_by(FuelSaleUploadLog.uploaded_at.desc()).limit(20).all()
    hist_routes, hist_airports = get_historical_rows()
    manual_routes = ManualRoute.query.order_by(ManualRoute.uploaded_at.desc()).all()
    airports = Airport.query.order_by(Airport.name.asc()).all()
    aircraft = Aircraft.query.order_by(Aircraft.name.asc()).all()
    flota_nombres = {f.get('nombre') for f in avion_model.get_flota().values() if f.get('nombre')}
    stats = {
        'routes': RouteMonthly.query.count() + len(hist_routes),
        'airports_monthly': AirportMonthly.query.count() + len(hist_airports),
        'airports_known': Airport.query.count(),
        'manual_routes': len(manual_routes),
    }
    admin_files = AdminFile.query.order_by(AdminFile.uploaded_at.desc()).all()
    return render_template('admin.html', logs=logs, airline_logs=airline_logs, fuel_sale_logs=fuel_sale_logs,
                            stats=stats, manual_routes=manual_routes, airports=airports, aircraft=aircraft,
                            flota_nombres=flota_nombres, admin_files=admin_files)


# ---------- Admin: Geocode (Airport coordinates) CRUD ----------

@app.route('/admin/airports', methods=['GET'])
def list_airports():
    if not session.get('is_admin'):
        return jsonify({"error": "No autorizado"}), 401
    airports = Airport.query.order_by(Airport.name.asc()).all()
    return jsonify([{"name": a.name, "lat": a.lat, "lon": a.lon, "is_argentina": a.is_argentina} for a in airports])


@app.route('/admin/airports/add', methods=['POST'])
def add_airport():
    if not session.get('is_admin'):
        return jsonify({"error": "No autorizado"}), 401

    data = request.get_json(silent=True) or request.form
    name = (data.get('name') or '').strip()
    lat = data.get('lat')
    lon = data.get('lon')
    is_argentina = str(data.get('is_argentina', '')).strip().lower() in ('1', 'true', 'on', 'yes')

    if not name:
        return jsonify({"error": "El nombre no puede estar vacío"}), 400
    try:
        lat = float(lat)
        lon = float(lon)
    except (TypeError, ValueError):
        return jsonify({"error": "Latitud/Longitud inválidas"}), 400
    if not (-90 <= lat <= 90) or not (-180 <= lon <= 180):
        return jsonify({"error": "Latitud/Longitud fuera de rango"}), 400

    existing = Airport.query.get(name)
    if existing:
        return jsonify({"error": f"Ya existe un aeropuerto llamado '{name}'. Usá editar en su lugar."}), 400

    db.session.add(Airport(name=name, lat=lat, lon=lon, is_argentina=is_argentina))
    db.session.commit()
    return jsonify({"ok": True, "name": name, "lat": lat, "lon": lon, "is_argentina": is_argentina})


@app.route('/admin/airports/update', methods=['POST'])
def update_airport():
    if not session.get('is_admin'):
        return jsonify({"error": "No autorizado"}), 401

    data = request.get_json(silent=True) or request.form
    original_name = (data.get('original_name') or '').strip()
    new_name = (data.get('name') or '').strip()
    lat = data.get('lat')
    lon = data.get('lon')
    is_argentina = str(data.get('is_argentina', '')).strip().lower() in ('1', 'true', 'on', 'yes')

    if not original_name:
        return jsonify({"error": "Falta el nombre original"}), 400
    airport = Airport.query.get(original_name)
    if not airport:
        return jsonify({"error": f"No se encontró el aeropuerto '{original_name}'"}), 404

    if not new_name:
        return jsonify({"error": "El nombre no puede estar vacío"}), 400
    try:
        lat = float(lat)
        lon = float(lon)
    except (TypeError, ValueError):
        return jsonify({"error": "Latitud/Longitud inválidas"}), 400
    if not (-90 <= lat <= 90) or not (-180 <= lon <= 180):
        return jsonify({"error": "Latitud/Longitud fuera de rango"}), 400

    if new_name != original_name:
        # Renaming: primary key changes, so need to check for collisions and
        # update any dependent rows (RouteMonthly, AirportMonthly, ManualRoute)
        if Airport.query.get(new_name):
            return jsonify({"error": f"Ya existe un aeropuerto llamado '{new_name}'"}), 400

        db.session.add(Airport(name=new_name, lat=lat, lon=lon, is_argentina=is_argentina))
        db.session.delete(airport)

        RouteMonthly.query.filter_by(origin=original_name).update({'origin': new_name})
        RouteMonthly.query.filter_by(dest=original_name).update({'dest': new_name})
        AirportMonthly.query.filter_by(airport=original_name).update({'airport': new_name})
        ManualRoute.query.filter_by(origin=original_name).update({'origin': new_name})
        ManualRoute.query.filter_by(dest=original_name).update({'dest': new_name})
    else:
        airport.lat = lat
        airport.lon = lon
        airport.is_argentina = is_argentina

    db.session.commit()
    return jsonify({"ok": True, "name": new_name, "lat": lat, "lon": lon, "is_argentina": is_argentina})


@app.route('/admin/airports/delete', methods=['POST'])
def delete_airport():
    if not session.get('is_admin'):
        return jsonify({"error": "No autorizado"}), 401

    data = request.get_json(silent=True) or request.form
    name = (data.get('name') or '').strip()
    if not name:
        return jsonify({"error": "Falta el nombre"}), 400

    airport = Airport.query.get(name)
    if not airport:
        return jsonify({"error": f"No se encontró el aeropuerto '{name}'"}), 404

    in_use = (
        RouteMonthly.query.filter((RouteMonthly.origin == name) | (RouteMonthly.dest == name)).first()
        or ManualRoute.query.filter((ManualRoute.origin == name) | (ManualRoute.dest == name)).first()
    )
    if in_use:
        return jsonify({"error": f"'{name}' está en uso por rutas existentes y no se puede borrar. "
                                  f"Editalo o borrá primero las rutas que lo usan."}), 400

    db.session.delete(airport)
    db.session.commit()
    return jsonify({"ok": True, "deleted": name})


# ---------- Admin: Aircraft catalog CRUD ----------

@app.route('/admin/airport_aliases', methods=['GET'])
def list_airport_aliases():
    if not session.get('is_admin'):
        return jsonify({"error": "No autorizado"}), 401
    aliases = AirportAlias.query.order_by(AirportAlias.alias.asc()).all()
    return jsonify([{"alias": a.alias, "airport_name": a.airport_name} for a in aliases])


@app.route('/admin/airport_aliases/add', methods=['POST'])
def add_airport_alias():
    if not session.get('is_admin'):
        return jsonify({"error": "No autorizado"}), 401

    data = request.get_json(silent=True) or request.form
    alias_raw = (data.get('alias') or '').strip()
    airport_name = (data.get('airport_name') or '').strip()

    if not alias_raw or not airport_name:
        return jsonify({"error": "Faltan datos (alias y aeropuerto)"}), 400
    if not Airport.query.get(airport_name):
        return jsonify({"error": f"'{airport_name}' no existe en Aeropuertos conocidos"}), 400

    alias_norm = _normalize_airport_name(alias_raw)
    existing = AirportAlias.query.get(alias_norm)
    if existing:
        existing.airport_name = airport_name
    else:
        db.session.add(AirportAlias(alias=alias_norm, airport_name=airport_name))
    db.session.commit()
    return jsonify({"ok": True, "alias": alias_norm, "airport_name": airport_name})


@app.route('/admin/airport_aliases/delete', methods=['POST'])
def delete_airport_alias():
    if not session.get('is_admin'):
        return jsonify({"error": "No autorizado"}), 401

    data = request.get_json(silent=True) or request.form
    alias_raw = (data.get('alias') or '').strip()
    if not alias_raw:
        return jsonify({"error": "Falta el alias"}), 400

    alias_norm = _normalize_airport_name(alias_raw)
    alias = AirportAlias.query.get(alias_norm)
    if not alias:
        return jsonify({"error": "No se encontró ese alias"}), 404
    db.session.delete(alias)
    db.session.commit()
    return jsonify({"ok": True, "deleted": alias_norm})


@app.route('/admin/fuel_sales_unmatched', methods=['GET'])
def fuel_sales_unmatched():
    """Nombres de AEROPUERTO ya guardados en FuelSale que no matchean (ni exacto, ni
    normalizado, ni por alias) con ningún Airport.name -- esos vuelos no van a aparecer
    en el mapa hasta agregar un alias o cargar el aeropuerto en 'Aeropuertos conocidos'."""
    if not session.get('is_admin'):
        return jsonify({"error": "No autorizado"}), 401

    known_norm = {_normalize_airport_name(a.name) for a in Airport.query.all()}
    alias_keys = {a.alias for a in AirportAlias.query.all()}

    distinct_aeropuertos = {row[0] for row in db.session.query(FuelSale.aeropuerto).distinct().all()}
    sin_match = sorted(
        a for a in distinct_aeropuertos
        if _normalize_airport_name(a) not in known_norm and _normalize_airport_name(a) not in alias_keys
    )
    return jsonify({"sin_match": sin_match})


@app.route('/admin/fuel_sales_unmatched/reprocess', methods=['POST'])
def reprocess_fuel_sales_airports():
    """Re-resuelve el AEROPUERTO de todas las filas de FuelSale ya guardadas contra los
    alias/aeropuertos conocidos actuales. Necesario después de agregar un alias nuevo,
    porque esas filas quedaron guardadas con el nombre crudo del Excel. Si al renombrar
    dos filas terminan compartiendo la misma clave (vuelo+aeropuerto+llaa+año+mes), se
    fusionan sumando cant_vuelos/volumen/ingreso en vez de duplicar."""
    if not session.get('is_admin'):
        return jsonify({"error": "No autorizado"}), 401

    known_airports_norm = {_normalize_airport_name(a.name): a.name for a in Airport.query.all()}
    alias_map = {a.alias: a.airport_name for a in AirportAlias.query.all()}

    renombrados, fusionados = 0, 0
    vistos = {}  # (vuelo, aeropuerto_canonico, llaa, anio, mes) -> objeto FuelSale ya procesado

    for r in FuelSale.query.all():
        canon, _ = resolve_airport_name(r.aeropuerto, known_airports_norm, alias_map)
        if not canon or canon == r.aeropuerto:
            key = (r.vuelo, r.aeropuerto, r.llaa, r.anio, r.mes)
            vistos.setdefault(key, r)
            continue

        key = (r.vuelo, canon, r.llaa, r.anio, r.mes)
        if key in vistos:
            destino_row = vistos[key]
            destino_row.cant_vuelos = (destino_row.cant_vuelos or 0) + (r.cant_vuelos or 0)
            destino_row.volumen_m3 = (destino_row.volumen_m3 or 0) + (r.volumen_m3 or 0)
            destino_row.ingreso_usd = (destino_row.ingreso_usd or 0) + (r.ingreso_usd or 0)
            db.session.delete(r)
            fusionados += 1
        else:
            r.aeropuerto = canon
            vistos[key] = r
            renombrados += 1

    db.session.commit()
    return jsonify({"ok": True, "renombrados": renombrados, "fusionados": fusionados})


@app.route('/admin/aircraft', methods=['GET'])
def list_aircraft():
    if not session.get('is_admin'):
        return jsonify({"error": "No autorizado"}), 401
    aircraft = Aircraft.query.order_by(Aircraft.name.asc()).all()
    return jsonify([{
        "name": a.name, "tipo_fuselaje": a.tipo_fuselaje,
        "consumo_hora_kg": a.consumo_hora_kg, "velocidad_crucero_kmh": a.velocidad_crucero_kmh,
        "asientos_default": a.asientos_default,
        "peso_operativo_kg": a.peso_operativo_kg,
        "elasticidad_override": a.elasticidad_override,
    } for a in aircraft])


def _validate_aircraft_fields(data):
    """Devuelve (name, tipo_fuselaje, consumo_hora_kg, velocidad_crucero_kmh, asientos_default,
    peso_operativo_kg, elasticidad_override, error). Los últimos dos son opcionales: si vienen
    vacíos, quedan en None (el frontend cae a la aproximación genérica por tipo_fuselaje)."""
    ERR = (None,) * 7
    name = (data.get('name') or '').strip()
    tipo_fuselaje = (data.get('tipo_fuselaje') or 'narrow').strip().lower()
    consumo_hora_kg = data.get('consumo_hora_kg')
    velocidad_crucero_kmh = data.get('velocidad_crucero_kmh')
    asientos_default = data.get('asientos_default')
    peso_operativo_kg = data.get('peso_operativo_kg')
    elasticidad_override = data.get('elasticidad_override')

    if not name:
        return ERR + ("El nombre no puede estar vacío",)
    if tipo_fuselaje not in ('regional', 'narrow', 'wide', 'piston'):
        return ERR + ("tipo_fuselaje debe ser 'regional', 'narrow', 'wide' o 'piston'",)
    try:
        consumo_hora_kg = float(consumo_hora_kg)
        velocidad_crucero_kmh = float(velocidad_crucero_kmh)
    except (TypeError, ValueError):
        return ERR + ("Consumo/hora y Velocidad crucero deben ser números",)
    if consumo_hora_kg <= 0 or velocidad_crucero_kmh <= 0:
        return ERR + ("Consumo/hora y Velocidad crucero deben ser mayores a 0",)

    asientos_int = None
    if asientos_default not in (None, ''):
        try:
            asientos_int = int(asientos_default)
        except (TypeError, ValueError):
            return ERR + ("Asientos por defecto debe ser un número entero",)

    peso_operativo_val = None
    if peso_operativo_kg not in (None, ''):
        try:
            peso_operativo_val = float(peso_operativo_kg)
        except (TypeError, ValueError):
            return ERR + ("Peso operativo (kg) debe ser un número",)
        if peso_operativo_val <= 0:
            return ERR + ("Peso operativo (kg) debe ser mayor a 0",)

    elasticidad_val = None
    if elasticidad_override not in (None, ''):
        try:
            elasticidad_val = float(elasticidad_override)
        except (TypeError, ValueError):
            return ERR + ("Elasticidad debe ser un número",)

    return (name, tipo_fuselaje, consumo_hora_kg, velocidad_crucero_kmh, asientos_int,
            peso_operativo_val, elasticidad_val, None)


@app.route('/admin/aircraft/add', methods=['POST'])
def add_aircraft():
    if not session.get('is_admin'):
        return jsonify({"error": "No autorizado"}), 401

    data = request.get_json(silent=True) or request.form
    (name, tipo_fuselaje, consumo, velocidad, asientos,
     peso_operativo_kg, elasticidad_override, error) = _validate_aircraft_fields(data)
    if error:
        return jsonify({"error": error}), 400

    if Aircraft.query.get(name):
        return jsonify({"error": f"Ya existe un avión llamado '{name}'. Usá editar en su lugar."}), 400

    db.session.add(Aircraft(name=name, tipo_fuselaje=tipo_fuselaje, consumo_hora_kg=consumo,
                             velocidad_crucero_kmh=velocidad, asientos_default=asientos,
                             peso_operativo_kg=peso_operativo_kg,
                             elasticidad_override=elasticidad_override))
    db.session.commit()
    return jsonify({"ok": True, "name": name})


@app.route('/admin/aircraft/update', methods=['POST'])
def update_aircraft():
    if not session.get('is_admin'):
        return jsonify({"error": "No autorizado"}), 401

    data = request.get_json(silent=True) or request.form
    original_name = (data.get('original_name') or '').strip()
    if not original_name:
        return jsonify({"error": "Falta el nombre original"}), 400
    aircraft = Aircraft.query.get(original_name)
    if not aircraft:
        return jsonify({"error": f"No se encontró el avión '{original_name}'"}), 404

    (new_name, tipo_fuselaje, consumo, velocidad, asientos,
     peso_operativo_kg, elasticidad_override, error) = _validate_aircraft_fields(data)
    if error:
        return jsonify({"error": error}), 400

    if new_name != original_name:
        if Aircraft.query.get(new_name):
            return jsonify({"error": f"Ya existe un avión llamado '{new_name}'"}), 400
        db.session.add(Aircraft(name=new_name, tipo_fuselaje=tipo_fuselaje, consumo_hora_kg=consumo,
                                 velocidad_crucero_kmh=velocidad, asientos_default=asientos,
                                 peso_operativo_kg=peso_operativo_kg,
                                 elasticidad_override=elasticidad_override))
        db.session.delete(aircraft)
        # Actualizar rutas manuales que usaban el nombre viejo, para no perder el vínculo al catálogo
        ManualRoute.query.filter_by(avion=original_name).update({'avion': new_name})
    else:
        aircraft.tipo_fuselaje = tipo_fuselaje
        aircraft.consumo_hora_kg = consumo
        aircraft.velocidad_crucero_kmh = velocidad
        aircraft.asientos_default = asientos
        aircraft.peso_operativo_kg = peso_operativo_kg
        aircraft.elasticidad_override = elasticidad_override

    db.session.commit()
    return jsonify({"ok": True, "name": new_name})


@app.route('/admin/aircraft/delete', methods=['POST'])
def delete_aircraft():
    if not session.get('is_admin'):
        return jsonify({"error": "No autorizado"}), 401

    data = request.get_json(silent=True) or request.form
    name = (data.get('name') or '').strip()
    if not name:
        return jsonify({"error": "Falta el nombre"}), 400

    aircraft = Aircraft.query.get(name)
    if not aircraft:
        return jsonify({"error": f"No se encontró el avión '{name}'"}), 404

    in_use = ManualRoute.query.filter_by(avion=name).first()
    if in_use:
        return jsonify({"error": f"'{name}' está en uso por rutas manuales existentes y no se puede "
                                  f"borrar. Cambiá el avión de esas rutas primero."}), 400

    db.session.delete(aircraft)
    db.session.commit()
    return jsonify({"ok": True, "deleted": name})


@app.route('/upload_manual', methods=['POST'])
@limiter.limit("10 per minute")
def upload_manual():
    if not session.get('is_admin'):
        return jsonify({"error": "No autorizado"}), 401

    f = request.files.get('file')
    if not f:
        return jsonify({"error": "No se recibió archivo"}), 400

    n_ok, errors = ingest_manual_routes(f)
    return jsonify({"ok": n_ok, "errors": errors})


@app.route('/upload_manual_single', methods=['POST'])
@limiter.limit("20 per minute")
def upload_manual_single():
    if not session.get('is_admin'):
        return jsonify({"error": "No autorizado"}), 401

    data = request.get_json(silent=True) or request.form
    partida = (data.get('partida') or '').strip()
    arribo = (data.get('arribo') or '').strip()
    tipo = (data.get('tipo') or 'internacional').strip().lower()
    avion = (data.get('avion') or '').strip()
    asientos = data.get('asientos')
    consumo_kg = data.get('consumo_kg')

    if not partida or not arribo:
        return jsonify({"error": "Partida y Arribo son obligatorios"}), 400
    if not avion:
        return jsonify({"error": "Elegí un avión"}), 400

    tipo_clean = 'cabotaje' if 'cab' in tipo else 'internacional'

    airports = {a.name: (a.lat, a.lon) for a in Airport.query.all()}
    unknown = [p for p in (partida, arribo) if p not in airports]
    if unknown:
        return jsonify({"error": f"Lugar(es) sin coordenadas: {', '.join(unknown)}. "
                                  f"Agregalos en 'Aeropuertos / Coordenadas' primero."}), 400

    aircraft_catalog = get_aircraft_catalog()
    aircraft_entry = aircraft_catalog.get(avion)
    if aircraft_entry is None:
        return jsonify({"error": f"'{avion}' no está en el catálogo de aviones"}), 400

    asientos_int = None
    if asientos not in (None, ''):
        try:
            asientos_int = int(asientos)
        except (TypeError, ValueError):
            return jsonify({"error": "Asientos inválido (no es un número)"}), 400
    else:
        asientos_int = Aircraft.query.get(avion).asientos_default

    consumo_kg_val = None
    if consumo_kg not in (None, ''):
        try:
            consumo_kg_val = float(consumo_kg)
        except (TypeError, ValueError):
            return jsonify({"error": "Consumo (kg) inválido (no es un número)"}), 400

    _upsert_single_manual_route(partida, arribo, tipo_clean, avion, asientos_int, consumo_kg_val,
                                 airports, aircraft_catalog)
    db.session.commit()
    return jsonify({"ok": True, "origin": partida, "dest": arribo})


@app.route('/manual_routes/delete/<int:route_id>', methods=['POST'])
def delete_manual_route(route_id):
    if not session.get('is_admin'):
        return jsonify({"error": "No autorizado"}), 401
    mr = ManualRoute.query.get_or_404(route_id)
    db.session.delete(mr)
    db.session.commit()
    return jsonify({"deleted": route_id})


@app.route('/admin/login', methods=['POST'])
@limiter.limit("5 per minute; 20 per hour")
def admin_login():
    if _passwords_match(request.form.get('password'), ADMIN_PASSWORD):
        session['is_admin'] = True
    return redirect(url_for('admin_page'))


@app.route('/admin/logout')
def admin_logout():
    session.pop('is_admin', None)
    return redirect(url_for('admin_page'))


@app.route('/upload', methods=['POST'])
@limiter.limit("10 per minute")
def upload():
    if not session.get('is_admin'):
        return jsonify({"error": "No autorizado"}), 401

    files = request.files.getlist('files')
    if not files:
        return jsonify({"error": "No se recibió ningún archivo"}), 400

    known_airports = {a.name for a in Airport.query.with_entities(Airport.name).all()}

    summary = []
    for f in files:
        if not f.filename or not f.filename.lower().endswith(('.xlsx', '.xlsm')):
            summary.append({"filename": f.filename, "error": "Formato no soportado (subí un Excel .xlsx/.xlsm)"})
            continue
        try:
            parsed = parse_workbook(f)
        except Exception as e:
            summary.append({"filename": f.filename, "error": str(e)})
            continue

        n_routes, n_airports, unknown_airports, descartadas = ingest_parsed_workbook(parsed, known_airports)

        nota_calidad = None
        if descartadas:
            ejemplos = ', '.join(f"{d['origin']}-{d['dest']} {d['year']}-{d['month']}"
                                 for d in descartadas[:5])
            nota_calidad = (f"{len(descartadas)} fila(s) descartadas por pax/vuelo "
                            f"físicamente imposible (dato corrompido en el archivo de ANAC, "
                            f"no un error de esta app): {ejemplos}"
                            + (', ...' if len(descartadas) > 5 else ''))

        db.session.add(UploadLog(
            filename=f.filename, rows_routes=n_routes, rows_airports=n_airports,
            note=" · ".join(filter(None, [
                ("Aeropuertos sin coordenadas (rutas omitidas): " + ", ".join(sorted(unknown_airports))
                 if unknown_airports else None),
                nota_calidad,
            ])) or None
        ))
        db.session.commit()

        summary.append({
            "filename": f.filename,
            "rows_routes": n_routes,
            "rows_airports": n_airports,
            "unknown_airports": sorted(unknown_airports),
            "descartadas_por_calidad": len(descartadas),
        })

        del parsed
        gc.collect()

    return jsonify({"summary": summary})


@app.route('/upload_aerolineas', methods=['POST'])
@limiter.limit("10 per minute")
def upload_aerolineas():
    if not session.get('is_admin'):
        return jsonify({"error": "No autorizado"}), 401

    try:
        f = request.files.get('file')
        if not f or not f.filename.lower().endswith('.pdf'):
            return jsonify({"error": "Subí un PDF válido (Informe Mensual de ANAC)"}), 400

        safe_name = secure_filename(f.filename) or 'upload.pdf'
        tmp_path = os.path.join('/tmp', safe_name)
        f.save(tmp_path)
        try:
            registros, lf_registros, warnings, (anio_informe, mes_informe) = extract_report(tmp_path)
        except Exception as e:
            return jsonify({"error": f"No se pudo parsear el PDF: {e}"}), 400
        finally:
            try:
                os.remove(tmp_path)
            except OSError:
                pass

        # Antes: una consulta SELECT por cada uno de los ~156 registros (ida y
        # vuelta a Postgres por cada uno). Ahora: UNA sola consulta que trae
        # de una todos los existentes para los años que trae este informe, y
        # después resolvemos insert/update en memoria con un dict. Esto no
        # solo es más rápido -- en un hosting con CPU limitada (Render free
        # tier), 156 round-trips secuenciales a la base pueden sumar varios
        # segundos y contribuir a superar el timeout del worker.
        anios_informe = {r['anio'] for r in registros}
        existentes = {
            (row.tipo, row.aerolinea, row.anio, row.mes): row
            for row in AirlineMonthly.query.filter(AirlineMonthly.anio.in_(anios_informe)).all()
        }

        insertados, actualizados = 0, 0
        for r in registros:
            existing = existentes.get((r['tipo'], r['aerolinea'], r['anio'], r['mes']))
            if existing:
                existing.vuelos = r['vuelos']
                existing.pax_000 = r['pax_000']
                existing.ocupacion = r['ocupacion']
                actualizados += 1
            else:
                db.session.add(AirlineMonthly(
                    tipo=r['tipo'], aerolinea=r['aerolinea'], anio=r['anio'], mes=r['mes'],
                    vuelos=r['vuelos'], pax_000=r['pax_000'], ocupacion=r['ocupacion'],
                ))
                insertados += 1

        existentes_lf = {
            (row.tipo, row.aerolinea, row.anio, row.mes): row
            for row in AirlineLoadFactorSnapshot.query.filter_by(anio=anio_informe).all()
        }

        lf_insertados, lf_actualizados = 0, 0
        for r in lf_registros:
            existing = existentes_lf.get((r['tipo'], r['aerolinea'], r['anio'], r['mes']))
            if existing:
                existing.lf_2025 = r['lf_2025']
                existing.lf_2026 = r['lf_2026']
                existing.variacion_pp = r['variacion_pp']
                lf_actualizados += 1
            else:
                db.session.add(AirlineLoadFactorSnapshot(
                    tipo=r['tipo'], aerolinea=r['aerolinea'], anio=r['anio'], mes=r['mes'],
                    lf_2025=r['lf_2025'], lf_2026=r['lf_2026'], variacion_pp=r['variacion_pp'],
                ))
                lf_insertados += 1

        db.session.add(AirlineUploadLog(
            filename=f.filename,
            registros_insertados=insertados + lf_insertados,
            registros_actualizados=actualizados + lf_actualizados,
            warnings='; '.join(warnings) if warnings else None,
        ))
        db.session.commit()

        return jsonify({
            "ok": True, "mes_informe": f"{mes_informe:02d}/{anio_informe}",
            "insertados": insertados, "actualizados": actualizados,
            "lf_insertados": lf_insertados, "lf_actualizados": lf_actualizados,
            "warnings": warnings,
        })
    except Exception as e:
        db.session.rollback()
        app.logger.exception("Error en /upload_aerolineas")
        return jsonify({"error": f"Error inesperado en el servidor: {e}"}), 500


@app.route('/aerolineas')
def aerolineas_page():
    if not (session.get('map_access') or session.get('is_admin')):
        return render_template('map_login.html')

    meses_disponibles = (
        db.session.query(AirlineMonthly.anio, AirlineMonthly.mes)
        .distinct()
        .order_by(AirlineMonthly.anio.desc(), AirlineMonthly.mes.desc())
        .all()
    )

    anio_sel = request.args.get('anio', type=int)
    mes_sel = request.args.get('mes', type=int)
    if not anio_sel or not mes_sel:
        if meses_disponibles:
            anio_sel, mes_sel = meses_disponibles[0]

    cabotaje = internacional = []
    total_cabotaje = total_internacional = None
    if anio_sel and mes_sel:
        cabotaje = (AirlineMonthly.query
                    .filter_by(tipo='cabotaje', anio=anio_sel, mes=mes_sel)
                    .filter(AirlineMonthly.aerolinea != 'TOTAL')
                    .order_by(AirlineMonthly.pax_000.desc())
                    .all())
        internacional = (AirlineMonthly.query
                          .filter_by(tipo='internacional', anio=anio_sel, mes=mes_sel)
                          .filter(AirlineMonthly.aerolinea != 'TOTAL')
                          .order_by(AirlineMonthly.pax_000.desc())
                          .all())
        total_cabotaje = AirlineMonthly.query.filter_by(
            tipo='cabotaje', anio=anio_sel, mes=mes_sel, aerolinea='TOTAL').first()
        total_internacional = AirlineMonthly.query.filter_by(
            tipo='internacional', anio=anio_sel, mes=mes_sel, aerolinea='TOTAL').first()

    return render_template(
        'aerolineas.html',
        meses_disponibles=meses_disponibles, anio_sel=anio_sel, mes_sel=mes_sel,
        cabotaje=cabotaje, internacional=internacional,
        total_cabotaje=total_cabotaje, total_internacional=total_internacional,
    )


@app.route('/admin/fuel_sales/borrar_todo', methods=['POST'])
@limiter.limit("5 per minute")
def admin_fuel_sales_borrar_todo():
    """Borra TODOS los registros de FuelSale y su historial de subidas — para poder cargar
    de cero un archivo con datos corregidos (precios, volúmenes) sin dejar mezclados datos
    viejos y nuevos del mismo período, que el upsert normal (por vuelo/aeropuerto/mes) no
    puede distinguir de una fila que simplemente ya no está en el archivo nuevo.

    Requiere admin Y una confirmación explícita en el body (`confirmar: "BORRAR"`) — no
    alcanza con estar logueado, porque es irreversible y afecta a toda la base, no a un
    archivo puntual."""
    if not session.get('is_admin'):
        return jsonify({"error": "No autorizado"}), 401
    try:
        data = request.get_json(silent=True) or {}
        if data.get('confirmar') != 'BORRAR':
            return jsonify({"error": "Falta la confirmación. Mandá {'confirmar': 'BORRAR'}."}), 400

        ventas_borradas = FuelSale.query.delete()
        logs_borrados = FuelSaleUploadLog.query.delete()
        db.session.commit()
        app.logger.warning(
            "Borrado masivo de FuelSale: %d ventas y %d logs eliminados (admin=%s)",
            ventas_borradas, logs_borrados, session.get('user_email') or session.get('user_name') or '?')

        return jsonify({"ok": True, "ventas_borradas": ventas_borradas, "logs_borrados": logs_borrados})
    except Exception as e:
        db.session.rollback()
        app.logger.exception("Error en /admin/fuel_sales/borrar_todo")
        return jsonify({"error": f"Error inesperado en el servidor: {e}"}), 500


# ---------- Admin: Archivos (carga/descarga momentánea de CSV/TXT/Excel) ----------

ALLOWED_ADMIN_FILE_EXT = ('.csv', '.txt', '.xlsx', '.xlsm')


@app.route('/admin/archivos/subir', methods=['POST'])
@limiter.limit("20 per minute")
def admin_archivos_subir():
    """Guarda cualquier CSV/TXT/Excel en la base tal cual, sin procesarlo ni leerlo -- es
    solo para poder bajarlo después desde otra máquina. Se guarda en Postgres (Neon) y no en
    el disco de Render, porque el filesystem del free tier es efímero y no sobrevive a un
    redeploy o restart."""
    if not session.get('is_admin'):
        return jsonify({"error": "No autorizado"}), 401

    files = request.files.getlist('files')
    if not files:
        return jsonify({"error": "No se recibió ningún archivo"}), 400

    subidos, errores = [], []
    for f in files:
        if not f.filename:
            continue
        if not f.filename.lower().endswith(ALLOWED_ADMIN_FILE_EXT):
            errores.append({"filename": f.filename, "error": "Solo se aceptan .csv, .txt, .xlsx, .xlsm"})
            continue
        safe_name = secure_filename(f.filename) or 'archivo'
        data = f.read()
        db.session.add(AdminFile(
            filename=safe_name,
            content_type=f.mimetype,
            size_bytes=len(data),
            content=data,
        ))
        subidos.append(safe_name)

    db.session.commit()
    return jsonify({"ok": True, "subidos": subidos, "errores": errores})


@app.route('/admin/archivos/<int:file_id>/descargar')
def admin_archivos_descargar(file_id):
    if not session.get('is_admin'):
        return jsonify({"error": "No autorizado"}), 401
    af = AdminFile.query.get_or_404(file_id)
    return send_file(
        io.BytesIO(af.content),
        mimetype=af.content_type or 'text/plain',
        as_attachment=True,
        download_name=af.filename,
    )


@app.route('/admin/archivos/<int:file_id>/borrar', methods=['POST'])
@limiter.limit("20 per minute")
def admin_archivos_borrar(file_id):
    if not session.get('is_admin'):
        return jsonify({"error": "No autorizado"}), 401
    af = AdminFile.query.get_or_404(file_id)
    db.session.delete(af)
    db.session.commit()
    return jsonify({"ok": True, "deleted": file_id})


@app.route('/admin/archivos/borrar_todo', methods=['POST'])
@limiter.limit("5 per minute")
def admin_archivos_borrar_todo():
    """Borra TODOS los archivos guardados en AdminFile. Pensado para uso momentáneo: subís,
    bajás cuando necesitás, y limpiás Neon de una para no dejar acumulando espacio."""
    if not session.get('is_admin'):
        return jsonify({"error": "No autorizado"}), 401
    data = request.get_json(silent=True) or {}
    if data.get('confirmar') != 'BORRAR':
        return jsonify({"error": "Falta la confirmación. Mandá {'confirmar': 'BORRAR'}."}), 400
    borrados = AdminFile.query.delete()
    db.session.commit()
    return jsonify({"ok": True, "borrados": borrados})


@app.route('/upload_fuel_sales', methods=['POST'])
@limiter.limit("10 per minute")
def upload_fuel_sales():
    if not session.get('is_admin'):
        return jsonify({"error": "No autorizado"}), 401

    try:
        f = request.files.get('file')
        if not f or not f.filename.lower().endswith(('.xlsx', '.xlsm')):
            return jsonify({"error": "Subí un archivo Excel (.xlsx) válido"}), 400

        try:
            registros = parse_fuel_sales_excel(f)
        except Exception as e:
            return jsonify({"error": f"No se pudo leer el Excel: {e}"}), 400

        if not registros:
            return jsonify({"error": "No se encontraron filas de datos en la hoja"}), 400

        sin_periodo = [r for r in registros if not r['mes'] or not r['anio']]
        if sin_periodo:
            return jsonify({"error": f"{len(sin_periodo)} fila(s) sin MES/AÑO válido. Revisá la hoja."}), 400

        # Resolver AEROPUERTO al nombre canónico que usa el mapa (tabla Airport).
        # El Excel suele venir en MAYÚSCULAS/sin tildes (EZEIZA) o con abreviaturas
        # distintas (COMODORO RIVADAVIA vs "Comod. Rivadavia" en el mapa); sin esto,
        # el click en el aeropuerto del mapa nunca encontraría estos datos.
        known_airports_norm = {_normalize_airport_name(a.name): a.name for a in Airport.query.all()}
        alias_map = {a.alias: a.airport_name for a in AirportAlias.query.all()}

        sin_match = set()
        for r in registros:
            canon, _ = resolve_airport_name(r['aeropuerto'], known_airports_norm, alias_map)
            if canon:
                r['aeropuerto'] = canon
            else:
                sin_match.add(r['aeropuerto'])  # se guarda tal cual vino; no aparecerá en el mapa hasta resolverse

        # Misma estrategia de batch upsert que /upload_aerolineas: una sola consulta
        # trayendo todo lo existente para los años del archivo, resuelto en memoria.
        anios = {r['anio'] for r in registros}
        existentes = {
            (row.vuelo, row.aeropuerto, row.llaa, row.anio, row.mes): row
            for row in FuelSale.query.filter(FuelSale.anio.in_(anios)).all()
        }

        insertados, actualizados = 0, 0
        for r in registros:
            key = (r['vuelo'], r['aeropuerto'], r['llaa'], r['anio'], r['mes'])
            existing = existentes.get(key)
            if existing:
                existing.cant_vuelos = r['cant_vuelos']
                existing.volumen_m3 = r['volumen_m3']
                existing.destino = r['destino']
                existing.precio_usd_l = r['precio_usd_l']
                existing.ingreso_usd = r['ingreso_usd']
                actualizados += 1
            else:
                db.session.add(FuelSale(**r))
                insertados += 1

        mes_ref, anio_ref = registros[0]['mes'], registros[0]['anio']
        warning_msg = None
        if sin_match:
            warning_msg = ("Aeropuertos sin match en el mapa (revisar en 'Alias de aeropuertos' o "
                            "agregarlos en Aeropuertos conocidos): " + ", ".join(sorted(sin_match)))

        db.session.add(FuelSaleUploadLog(
            filename=f.filename, mes=mes_ref, anio=anio_ref,
            registros_insertados=insertados, registros_actualizados=actualizados,
            warnings=warning_msg,
        ))
        db.session.commit()

        return jsonify({
            "ok": True, "periodo": f"{mes_ref:02d}/{anio_ref}",
            "insertados": insertados, "actualizados": actualizados,
            "sin_match": sorted(sin_match),
        })
    except Exception as e:
        db.session.rollback()
        app.logger.exception("Error en /upload_fuel_sales")
        return jsonify({"error": f"Error inesperado en el servidor: {e}"}), 500


@app.route('/api/fuel_sales_diagnostico')
@limiter.limit("20 per minute")
def api_fuel_sales_diagnostico():
    """Para cualquiera con acceso al mapa (no hace falta ser admin): dos cosas distintas
    sobre los despachos de combustible cargados.

    1) Cuántos registros de FuelSale, para el período que se esté mirando, tienen un
       AEROPUERTO que no matchea con ningún Airport conocido -- esos vuelos existen en la
       base, se subieron bien, pero nunca van a dibujarse en el mapa ni sumar al aeropuerto
       que corresponde, hasta que se les resuelva un alias.

    2) El total de registros y volumen que existen para ese período SIN IMPORTAR si
       matchearon o no, más la lista completa de qué períodos (año-mes) tienen algo cargado.
       Esto responde una pregunta distinta y más básica: "¿el Excel que subí realmente
       quedó guardado con año/mes 2026-07, o se guardó con otro período por error?" — el
       punto "Despacho YPF real" del gráfico de Proyecciones sale de exactamente esta suma,
       sin filtrar por aeropuerto, así que si no aparece para un mes puntual, lo más
       probable es que sencillamente no haya ninguna fila con ese año/mes en la base
       (revisar el período que se le asignó al subir el Excel), o que el navegador esté
       mostrando una carga vieja de antes de la subida (conviene refrescar la página).

    Sin este endpoint, "no veo todos los aeropuertos de julio" o "falta el despacho de
    julio" son indistinguibles entre "no se subió nada de julio" (nada que arreglar acá) y
    "se subió pero bajo un período o nombre distinto" (se corrige sin volver a subir el
    archivo)."""
    if not session.get('map_access'):
        return jsonify({"error": "No autorizado"}), 401

    anio = request.args.get('anio', type=int)
    mes = request.args.get('mes', type=int)

    known_norm = {_normalize_airport_name(a.name) for a in Airport.query.all()}
    alias_keys = {a.alias for a in AirportAlias.query.all()}

    q = FuelSale.query
    if anio:
        q = q.filter_by(anio=anio)
    if mes:
        q = q.filter_by(mes=mes)
    filas = q.all()

    sin_match = {}
    for r in filas:
        norm = _normalize_airport_name(r.aeropuerto)
        if norm not in known_norm and norm not in alias_keys:
            e = sin_match.setdefault(r.aeropuerto, {'registros': 0, 'volumen_m3': 0.0})
            e['registros'] += 1
            e['volumen_m3'] += r.volumen_m3 or 0.0

    aeropuertos = [{'aeropuerto': k, 'registros': v['registros'],
                    'volumen_m3': round(v['volumen_m3'], 3)}
                   for k, v in sorted(sin_match.items())]

    periodos = sorted({(r.anio, r.mes) for r in
                       db.session.query(FuelSale.anio, FuelSale.mes).distinct().all()})

    return jsonify({
        "anio": anio, "mes": mes,
        # Todo lo que hay cargado para el período pedido, matchee o no el aeropuerto — esto
        # es exactamente lo que suma el punto "Despacho YPF real" del gráfico de Proyecciones.
        "total_registros_periodo": len(filas),
        "total_volumen_periodo_m3": round(sum(r.volumen_m3 or 0.0 for r in filas), 3),
        "aeropuertos_sin_match": aeropuertos,
        "total_registros_sin_match": sum(a['registros'] for a in aeropuertos),
        "total_volumen_sin_match_m3": round(sum(a['volumen_m3'] for a in aeropuertos), 3),
        # Todos los períodos que tienen AL MENOS un registro cargado, para poder confirmar
        # de un vistazo si el período que se esperaba (ej. 2026-7) realmente está ahí.
        "periodos_con_datos": [{"anio": a, "mes": m} for a, m in periodos],
    })


@app.route('/api/fuel_sales')
@limiter.limit("30 per minute")
def api_fuel_sales():
    if not session.get('map_access'):
        return jsonify({"error": "No autorizado"}), 401

    # Segunda capa de contraseña: solo protege ingreso/precio (los USD). Vuelos, despachos y
    # volumen se devuelven siempre, sin necesidad de desbloquear.
    fuel_access = bool(session.get('fuel_access'))

    aeropuerto = (request.args.get('aeropuerto') or '').strip()
    if not aeropuerto:
        return jsonify({"error": "Falta el parámetro aeropuerto"}), 400

    # Períodos vienen como p=YYYY-M (uno por cada combinación año/mes tildada en el mapa).
    # Si no se manda ninguno, se devuelven todos los períodos cargados para ese aeropuerto.
    periods = set()
    for p in request.args.getlist('p'):
        try:
            y, m = p.split('-')
            periods.add((int(y), int(m)))
        except (ValueError, AttributeError):
            continue

    query = FuelSale.query.filter_by(aeropuerto=aeropuerto)
    if periods:
        anios = {y for y, m in periods}
        meses = {m for y, m in periods}
        query = query.filter(FuelSale.anio.in_(anios), FuelSale.mes.in_(meses))

    rows = query.all()
    if periods:
        rows = [r for r in rows if (r.anio, r.mes) in periods]

    grouped = {}
    for r in rows:
        key = (r.vuelo, r.llaa)
        g = grouped.setdefault(key, {
            'vuelo': r.vuelo, 'llaa': r.llaa, 'destino': r.destino,
            'cant_vuelos': 0, 'volumen_m3': 0.0, 'ingreso_usd': 0.0,
        })
        g['cant_vuelos'] += r.cant_vuelos or 0
        g['volumen_m3'] += r.volumen_m3 or 0.0
        g['ingreso_usd'] += r.ingreso_usd or 0.0
        if not g['destino'] and r.destino:
            g['destino'] = r.destino

    result_rows = []
    for g in grouped.values():
        litros = (g['volumen_m3'] or 0) * 1000
        precio_prom = (g['ingreso_usd'] / litros) if litros else None
        ingreso_por_vuelo = round(g['ingreso_usd'] / g['cant_vuelos'], 2) if g['cant_vuelos'] else 0
        row_out = {
            'vuelo': g['vuelo'], 'llaa': g['llaa'], 'destino': g['destino'],
            'cant_vuelos': g['cant_vuelos'],
            'volumen_m3': g['volumen_m3'],
            'volumen_por_vuelo': round(g['volumen_m3'] / g['cant_vuelos'], 3) if g['cant_vuelos'] else 0,
            # Campos detrás de la segunda contraseña: null si la sesión no la desbloqueó.
            'precio_usd_l': (round(precio_prom, 6) if precio_prom is not None else None) if fuel_access else None,
            'ingreso_usd': g['ingreso_usd'] if fuel_access else None,
            'ingreso_por_vuelo': ingreso_por_vuelo if fuel_access else None,
        }
        result_rows.append(row_out)

    # Sin desbloquear no hay ingreso_usd para ordenar; se ordena por volumen en su lugar.
    if fuel_access:
        result_rows.sort(key=lambda r: r['ingreso_usd'], reverse=True)
    else:
        result_rows.sort(key=lambda r: r['volumen_m3'], reverse=True)
    llaa_list = sorted({r['llaa'] for r in result_rows if r['llaa']})

    return jsonify({
        "aeropuerto": aeropuerto,
        "rows": result_rows,
        "llaa_list": llaa_list,
        "fuel_access": fuel_access,
        "total_ingreso": round(sum(r['ingreso_usd'] for r in result_rows), 2) if fuel_access else None,
        "total_volumen": round(sum(r['volumen_m3'] for r in result_rows), 3),
    })


# ============================================================================
# Consumo nacional de Jet A-1 y proyecciones
# ============================================================================

# Cuánto del combustible de un vuelo se carga EN ARGENTINA.
#
# ANAC informa cada par origen-destino una sola vez, con los vuelos de los dos sentidos
# sumados. Entonces:
#   - Cabotaje: los dos tramos despegan de un aeropuerto argentino, así que el 100% del
#     combustible de esos vuelos se carga acá.  -> factor 1.0
#   - Internacional: solo el tramo que sale de Argentina reposta acá; el de vuelta carga
#     en el exterior. Como los vuelos vienen sumados, se toma la mitad. -> factor 0.5
#
# El 0.5 es un piso conservador: no contempla tankering (cargar de más acá cuando el
# combustible está más barato) ni el hecho de que las salidas internacionales embarcan
# además combustible de contingencia y alternativa. Se deja como constante editable en vez
# de hardcodearlo suelto en el cálculo.
FACTOR_UPLIFT_CABOTAJE = 1.0
FACTOR_UPLIFT_INTERNACIONAL = 0.5

_MES_A_NUM = {m: i + 1 for i, m in enumerate(MONTH_ORDER)}


def _serie_consumo_nacional(aeropuerto=None):
    """Serie mensual de Jet A-1 cargado en Argentina, en m³, a partir del tráfico real.

    Devuelve (serie, detalle_rutas) donde serie es una lista ordenada de dicts por mes.
    El avión de cada ruta se elige mes a mes según la ocupación de ese mes: es justamente
    lo que hace que el total sea sensible al mix de flota y no solo al número de vuelos.

    Si se pasa `aeropuerto`, la serie queda acotada a lo que se carga EN ESE aeropuerto en
    particular, no el total del país:
    - Cabotaje (los dos extremos son argentinos): la ruta ya factura el 100% del combustible
      porque los dos tramos cargan en Argentina, pero repartido entre DOS aeropuertos. Si el
      aeropuerto elegido es uno de los dos extremos, le toca la mitad del total de la ruta.
    - Internacional (un solo extremo argentino): el aeropuerto elegido tiene que ser
      justamente ese extremo — ahí le toca el 100% de lo que la ruta aporta al total
      nacional (que ya viene descontado al 50% por FACTOR_UPLIFT_INTERNACIONAL, así que no
      se vuelve a dividir)."""
    hist_routes, _ = get_historical_rows()
    db_routes = [
        {'tipo': r.tipo, 'origin': r.origin, 'dest': r.dest, 'year': r.year,
         'month': r.month, 'vuelos': r.vuelos, 'pax': r.pax}
        for r in RouteMonthly.query.all()
    ]

    _all_airports = Airport.query.all()
    airports = {a.name: (a.lat, a.lon) for a in _all_airports}
    es_argentino = {a.name: a.is_argentina for a in _all_airports}

    # Las rutas manuales del admin tienen avión y consumo confirmados: registrarlas para
    # que get_aircraft_info las respete también acá.
    aircraft_catalog = get_aircraft_catalog()
    for mr in ManualRoute.query.all():
        avion_model.register_manual_route(
            mr.origin, mr.dest, mr.tipo, mr.avion, mr.asientos, airports,
            consumo_kg_override=mr.consumo_kg_manual, aircraft_catalog=aircraft_catalog
        )

    acc = {}       # (anio, mes) -> dict con los totales
    por_ruta = {}  # (o, d, tipo) -> m³ acumulados (para el ranking)

    for r in hist_routes + db_routes:
        o, d = r['origin'], r['dest']
        if o not in airports or d not in airports:
            continue
        vuelos = r['vuelos'] or 0
        if vuelos <= 0:
            continue

        o_ar = es_argentino.get(o, False)
        d_ar = es_argentino.get(d, False)
        if not o_ar and not d_ar:
            continue  # ruta que no toca Argentina: no carga combustible acá
        es_cabotaje = o_ar and d_ar
        factor = FACTOR_UPLIFT_CABOTAJE if es_cabotaje else FACTOR_UPLIFT_INTERNACIONAL

        if aeropuerto:
            if aeropuerto not in (o, d):
                continue  # esta ruta no toca el aeropuerto pedido: no aporta a su serie
            if es_cabotaje:
                factor *= 0.5  # los dos extremos son argentinos: se reparte por mitades

        pax = r['pax'] or 0
        ppv = (pax / vuelos) if (pax and vuelos) else None
        info = avion_model.get_aircraft_info(o, d, r['tipo'], airports, pax_por_vuelo=ppv)
        if not info or not info.get('consumo_total_m3'):
            continue

        m3 = info['consumo_total_m3'] * vuelos * factor
        mes_num = _MES_A_NUM.get(r['month'])
        if not mes_num:
            continue
        try:
            anio = int(r['year'])
        except (TypeError, ValueError):
            continue

        e = acc.setdefault((anio, mes_num), {
            'cabotaje_m3': 0.0, 'internacional_m3': 0.0,
            'vuelos_cabotaje': 0, 'vuelos_intl': 0, 'pax': 0,
        })
        if es_cabotaje:
            e['cabotaje_m3'] += m3
            e['vuelos_cabotaje'] += vuelos
        else:
            e['internacional_m3'] += m3
            e['vuelos_intl'] += vuelos
        e['pax'] += pax

        pr = por_ruta.setdefault((o, d, r['tipo']), {'m3': 0.0, 'vuelos': 0,
                                                     'avion': info['avion']})
        pr['m3'] += m3
        pr['vuelos'] += vuelos

    serie = []
    for (anio, mes), e in sorted(acc.items()):
        total = e['cabotaje_m3'] + e['internacional_m3']
        serie.append({
            'anio': anio, 'mes': mes,
            'ym': f"{anio}-{MONTH_ORDER[mes - 1]}",
            'cabotaje_m3': round(e['cabotaje_m3'], 1),
            'internacional_m3': round(e['internacional_m3'], 1),
            'total_m3': round(total, 1),
            'total_t': round(total * avion_model.JETA1_KG_PER_M3 / 1000.0, 1),
            'vuelos': e['vuelos_cabotaje'] + e['vuelos_intl'],
            'pax': e['pax'],
        })

    # Marcar los meses que caen dentro de alguna ventana de exclusión activa, para que el
    # front pueda dejarlos afuera del cálculo de estacionalidad y del nivel base sin perder
    # el dato: siguen viéndose en el gráfico, solo no pesan en la proyección. Un evento sin
    # aeropuerto (None) aplica siempre; uno con aeropuerto puntual solo aplica cuando se está
    # mirando justo esa aeroplanta.
    exclusiones = ProyeccionExclusion.query.filter_by(activo=True).all()
    exclusiones = [ex for ex in exclusiones if not ex.aeropuerto or ex.aeropuerto == aeropuerto]
    if exclusiones:
        for s in serie:
            idx = s['anio'] * 12 + s['mes']
            motivos = [ex.motivo for ex in exclusiones
                      if ex.desde_anio * 12 + ex.desde_mes <= idx <= ex.hasta_anio * 12 + ex.hasta_mes]
            s['excluido'] = bool(motivos)
            s['motivo_exclusion'] = '; '.join(motivos) if motivos else None
    else:
        for s in serie:
            s['excluido'] = False
            s['motivo_exclusion'] = None

    ranking = sorted(
        ({'origin': k[0], 'dest': k[1], 'tipo': k[2], 'm3': round(v['m3'], 1),
          'vuelos': v['vuelos'], 'avion': v['avion']}
         for k, v in por_ruta.items()),
        key=lambda x: -x['m3']
    )[:40]

    return serie, ranking


@app.route('/api/consumo_nacional')
@limiter.limit("20 per minute")
def api_consumo_nacional():
    """Serie histórica mensual del Jet A-1 cargado en Argentina, estimado desde el tráfico
    de ANAC más el modelo de asignación de avión por ocupación.

    Con ?aeropuerto=NOMBRE, la serie queda acotada a esa aeroplanta puntual en vez del total
    del país (ver el reparto por mitades en _serie_consumo_nacional). Solo admite
    aeropuertos argentinos — no tiene sentido pedir "lo que carga" un aeropuerto extranjero.

    Devuelve también, mes a mes, el volumen REALMENTE despachado que haya cargado en
    FuelSale. Eso permite calcular un factor de calibración: el modelo es bottom-up y
    arrastra los sesgos de sus supuestos (tankering, combustible de contingencia, mix de
    flota), así que contrastarlo contra el despacho real es la única forma de saber cuánto
    se le está errando. Ojo con la interpretación: FuelSale son los despachos de YPF, no el
    total del mercado argentino, así que el cociente es un share, no un error."""
    if not session.get('map_access'):
        return jsonify({"error": "No autorizado"}), 401

    aeropuerto = (request.args.get('aeropuerto') or '').strip() or None
    if aeropuerto:
        ok = db.session.query(Airport.name).filter(
            Airport.name == aeropuerto, Airport.is_argentina.is_(True)).first()
        if not ok:
            return jsonify({"error": f"'{aeropuerto}' no es un aeropuerto argentino conocido"}), 400

    try:
        serie, ranking = _serie_consumo_nacional(aeropuerto=aeropuerto)

        # Volumen real despachado por mes. Sin aeropuerto, se suman todos; con aeropuerto,
        # solo lo que FuelSale tiene cargado para ESE aeropuerto puntual (son los despachos
        # reales de YPF ahí, no hace falta ningún reparto — ya vienen por aeropuerto).
        query = db.session.query(FuelSale.anio, FuelSale.mes, func.sum(FuelSale.volumen_m3))
        if aeropuerto:
            query = query.filter(FuelSale.aeropuerto == aeropuerto)
        reales = {}
        for anio, mes, vol in query.group_by(FuelSale.anio, FuelSale.mes).all():
            if mes and 1 <= mes <= 12:
                reales[(anio, mes)] = float(vol or 0.0)

        # NO se crean filas sintéticas para los meses que tienen despacho pero todavía no
        # tienen tráfico de ANAC. `serie` es el histórico REAL y solo eso: un mes sin rutas
        # de ANAC no es histórico, es un mes a proyectar. Si se metiera igual, el frontend lo
        # tomaría como "último mes conocido" y arrancaría la proyección después, dejándolo
        # dibujado en cero.
        #
        # En su lugar, los datos reales de todos los períodos viajan aparte en
        # `reales_por_periodo`: el frontend los engancha a la fila que corresponda, sea del
        # histórico o de la proyección. Así julio aparece como UNA sola fila proyectada, con
        # su despacho real al lado, en vez de dos filas separadas.
        # Mercado real por empresa (planilla propia). Solo tiene sentido a nivel país: es un
        # total de mercado, no está desagregado por aeroplanta, así que con una aeroplanta
        # elegida se omite en vez de comparar peras con manzanas.
        mercado_all = ({(m.anio, m.mes): m for m in MercadoMensual.query.all()}
                       if not aeropuerto else {})

        def _despacho_efectivo(anio, mes):
            """Despacho de YPF del mes, con respaldo. Devuelve (m3, fuente).

            Prioridad al Excel de despachos (FuelSale): es el dato propio, desagregado por
            aeropuerto y por vuelo. Si ese mes todavía no se subió, se cae a la columna YPF
            de la planilla de mercado, que es la misma magnitud a nivel país y así no queda
            un hueco en la serie solo por el orden en que llegan los archivos.

            El respaldo NO aplica cuando se filtra por aeroplanta: la planilla de mercado es
            un total nacional, no se puede repartir entre aeroplantas. Por eso, al elegir una
            aeroplanta, solo se ven los meses que tengan FuelSale cargado."""
            real = reales.get((anio, mes))
            if real is not None:
                return real, 'fuel_sales'
            m = mercado_all.get((anio, mes))
            if m and m.ypf_m3:
                return float(m.ypf_m3), 'mercado'
            return None, None

        for s in serie:
            real, fuente = _despacho_efectivo(s['anio'], s['mes'])
            s['real_m3'] = round(real, 1) if real is not None else None
            s['real_m3_fuente'] = fuente
            s['ratio_real_modelo'] = (round(real / s['total_m3'], 4)
                                      if (real and s['total_m3']) else None)

            m = mercado_all.get((s['anio'], s['mes']))
            if not m:
                s['mercado_total_m3'] = None
                s['mercado_ypf_m3'] = None
                s['desvio_modelo_pct'] = None
                s['share_ypf_real'] = None
                continue
            s['mercado_total_m3'] = round(m.total_m3, 1)
            s['mercado_ypf_m3'] = round(m.ypf_m3, 1) if m.ypf_m3 else None
            # ESTE es el error real del modelo: modelo contra el total del mercado.
            # (ratio_real_modelo, en cambio, es el share de YPF, no un error.)
            s['desvio_modelo_pct'] = (round((s['total_m3'] / m.total_m3 - 1) * 100, 2)
                                      if (s['total_m3'] and m.total_m3) else None)
            s['share_ypf_real'] = (round(m.ypf_m3 / m.total_m3, 4)
                                   if (m.ypf_m3 and m.total_m3) else None)

        # Datos reales de TODOS los períodos, incluidos los que aún no tienen tráfico de ANAC
        # y por lo tanto no están en `serie`. El frontend los engancha a la fila que
        # corresponda: si el mes ya es histórico, a esa fila; si todavía se proyecta (julio,
        # con el Excel de combustible subido pero el de rutas no), a la fila proyectada. Así
        # el despacho real de un mes proyectado se ve igual, en una sola fila.
        periodos = set(reales) | set(mercado_all)
        reales_por_periodo = {}
        for (anio, mes) in periodos:
            m = mercado_all.get((anio, mes))
            real, fuente = _despacho_efectivo(anio, mes)
            reales_por_periodo[f"{anio}-{mes}"] = {
                'real_m3': round(real, 1) if real is not None else None,
                'real_m3_fuente': fuente,
                'mercado_total_m3': round(m.total_m3, 1) if m else None,
                'mercado_ypf_m3': round(m.ypf_m3, 1) if (m and m.ypf_m3) else None,
                'share_ypf_real': (round(m.ypf_m3 / m.total_m3, 4)
                                   if (m and m.ypf_m3 and m.total_m3) else None),
            }

        return jsonify({
            "serie": serie,
            "ranking_rutas": ranking,
            "aeropuerto": aeropuerto,
            "reales_por_periodo": reales_por_periodo,
            "factores": {
                "cabotaje": FACTOR_UPLIFT_CABOTAJE,
                "internacional": FACTOR_UPLIFT_INTERNACIONAL,
            },
            "meses_con_dato_real": sum(1 for s in serie if s['real_m3']),
            "month_order": MONTH_ORDER,
        })
    except Exception as e:
        app.logger.exception("Error en /api/consumo_nacional")
        return jsonify({"error": str(e), "type": type(e).__name__}), 500


@app.route('/api/estimar_ruta')
@limiter.limit("60 per minute")
def api_estimar_ruta():
    """Estima avión y consumo de una ruta arbitraria, para el previsualizador de la
    pestaña de proyecciones. No guarda nada."""
    if not session.get('map_access'):
        return jsonify({"error": "No autorizado"}), 401

    origin = (request.args.get('origin') or '').strip()
    dest = (request.args.get('dest') or '').strip()
    if not origin or not dest:
        return jsonify({"error": "Faltan origen y/o destino"}), 400
    if origin == dest:
        return jsonify({"error": "El origen y el destino no pueden ser el mismo"}), 400

    try:
        ppv = float(request.args.get('pax_por_vuelo') or 0) or None
    except ValueError:
        ppv = None
    avion_forzado = (request.args.get('avion') or '').strip() or None

    _all_airports = Airport.query.all()
    airports = {a.name: (a.lat, a.lon) for a in _all_airports}
    es_argentino = {a.name: a.is_argentina for a in _all_airports}

    if origin not in airports or dest not in airports:
        faltan = [x for x in (origin, dest) if x not in airports]
        return jsonify({"error": f"Sin coordenadas para: {', '.join(faltan)}"}), 400

    o_ar = es_argentino.get(origin, False)
    d_ar = es_argentino.get(dest, False)
    if not o_ar and not d_ar:
        return jsonify({"error": "La ruta no toca Argentina: no carga combustible acá"}), 400
    tipo = 'cabotaje' if (o_ar and d_ar) else 'internacional'
    factor = FACTOR_UPLIFT_CABOTAJE if tipo == 'cabotaje' else FACTOR_UPLIFT_INTERNACIONAL

    dist = avion_model.haversine(*airports[origin], *airports[dest])

    if avion_forzado and avion_forzado in avion_model.get_flota():
        key = avion_model._key(origin, dest)
        tons, origen_dato = avion_model.consumo_toneladas(avion_forzado, dist, key)
        ficha = avion_model.get_flota()[avion_forzado]
        fac, lf = avion_model.ajuste_por_ocupacion(avion_forzado, ppv)
        kg = (tons or 0.0) * 1000.0 * fac
        info = {
            'avion': ficha['nombre'], 'avion_codigo': avion_forzado,
            'asientos': ficha['asientos'], 'distancia_km': round(dist, 1),
            'consumo_total_kg': round(kg, 1),
            'consumo_total_m3': round(kg / avion_model.JETA1_KG_PER_M3, 3),
            'load_factor': round(lf, 3) if lf is not None else None,
            'fuente': ('Real (planilla de consumo YPF)' if origen_dato == 'real'
                       else 'Estimado (curva calibrada del tipo)'),
            'dato_consumo': origen_dato,
        }
    else:
        info = avion_model.get_aircraft_info(origin, dest, tipo, airports, pax_por_vuelo=ppv)
    if not info:
        return jsonify({"error": "No se pudo estimar el consumo de esa ruta"}), 400

    return jsonify({
        "origin": origin, "dest": dest, "tipo": tipo,
        "factor_uplift": factor,
        "info": info,
        "opciones": avion_model.opciones_para_rango(origin, dest, airports,
                                                    pax_min=ppv, pax_max=ppv, dist=dist),
    })


# ---------- Escenarios de proyección (accesibles con la clave del mapa, sin admin) ----------

def _proyeccion_config(escenario):
    cfg = ProyeccionConfig.query.get(escenario)
    if cfg is None:
        cfg = ProyeccionConfig(escenario=escenario)
        db.session.add(cfg)
        db.session.commit()
    return cfg


def _ruta_proyeccion_to_dict(r):
    return {
        'id': r.id, 'escenario': r.escenario, 'origin': r.origin, 'dest': r.dest,
        'tipo': r.tipo, 'vuelos_mes': r.vuelos_mes, 'pax_por_vuelo': r.pax_por_vuelo,
        'avion_forzado': r.avion_forzado, 'signo': r.signo,
        'desde_anio': r.desde_anio, 'desde_mes': r.desde_mes,
        'hasta_anio': r.hasta_anio, 'hasta_mes': r.hasta_mes,
        'activo': r.activo, 'nota': r.nota, 'creado_por': r.creado_por,
    }


@app.route('/api/deploy_status')
def api_deploy_status():
    """Diagnóstico de deploy, pensado para pegar la URL en el navegador y ver en 2 segundos
    qué archivos del modelo llegaron al servidor, sin entrar al dashboard de Render a leer
    logs. Requiere estar logueado al mapa, nada más — no expone datos sensibles."""
    if not session.get('map_access'):
        return jsonify({"error": "No autorizado"}), 401

    base = os.path.dirname(os.path.abspath(__file__))
    archivos = {
        'flota.json': os.path.exists(os.path.join(base, 'flota.json')),
        'consumo_rutas.json': os.path.exists(os.path.join(base, 'consumo_rutas.json')),
        'templates/proyecciones.html': os.path.exists(
            os.path.join(base, 'templates', 'proyecciones.html')),
        'geocode.py (con Orán)': 'Orán' in COORDS,
    }
    # También hay que confirmar que lo que se cargó tiene contenido de verdad, no un archivo
    # vacío o con el JSON de otro (como pasó con flota.json duplicado en consumo_rutas.json
    # en un deploy anterior).
    tipos_flota = len(avion_model.get_flota())
    rutas_consumo = len(avion_model.get_matriz())

    todo_ok = (all(archivos.values()) and tipos_flota > 0 and rutas_consumo > 0)

    return jsonify({
        "todo_ok": todo_ok,
        "version_esperada": MODEL_VERSION,
        "archivos_encontrados": archivos,
        "tipos_de_avion_cargados": tipos_flota,
        "rutas_con_consumo_real_cargadas": rutas_consumo,
        "diagnostico": (
            "Todo en orden." if todo_ok else
            "Falta o está vacío alguno de los archivos del modelo — revisá 'archivos_encontrados' "
            "y comparalo contra LEEME_deploy.md. Si un archivo da 'true' pero los tipos/rutas dan "
            "0, el archivo existe pero está vacío o con el contenido equivocado."
        ),
    })


@app.route('/proyecciones')
def proyecciones_page():
    if not session.get('map_access'):
        return render_template('map_login.html')
    try:
        return render_template('proyecciones.html')
    except Exception as e:
        # Un TemplateNotFound acá casi siempre es un problema de deploy, no de código: el
        # archivo templates/proyecciones.html no llegó con ese nombre exacto (extensión
        # perdida, mayúscula de más, carpeta equivocada). Sin este bloque, el usuario ve la
        # página genérica "Internal Server Error" de Werkzeug y no tiene forma de saber por
        # qué sin entrar al dashboard de Render a leer logs.
        app.logger.exception("Error renderizando /proyecciones")
        return f"""
        <div style="font-family:Arial,sans-serif; max-width:640px; margin:60px auto;
                    background:#10162e; color:#eee; padding:28px 32px; border-radius:12px;
                    border:1px solid #2a3358;">
          <h2 style="color:#ff9a9a; margin-top:0;">No se pudo cargar Proyecciones</h2>
          <p style="color:#c8cfe6;">Error: <code>{type(e).__name__}: {e}</code></p>
          <p style="color:#9aa4c7; font-size:14px; line-height:1.6;">
            Esto casi siempre significa que <code>templates/proyecciones.html</code> no llegó
            al deploy con ese nombre exacto (falta la extensión <code>.html</code>, quedó con
            mayúsculas, o en la carpeta equivocada — Linux distingue mayúsculas de minúsculas).
            Visitá <a href="/api/deploy_status" style="color:#8fb2ff;">/api/deploy_status</a>
            para ver qué archivos encuentra el servidor.
          </p>
          <p><a href="/" style="color:#8fb2ff;">← Volver al mapa</a></p>
        </div>
        """, 500


@app.route('/api/proyeccion/escenarios')
@limiter.limit("60 per minute")
def api_proyeccion_escenarios():
    if not session.get('map_access'):
        return jsonify({"error": "No autorizado"}), 401
    nombres = sorted({e for (e,) in db.session.query(ProyeccionRuta.escenario).distinct()} |
                     {c.escenario for c in ProyeccionConfig.query.all()} | {'Base'})
    return jsonify({"escenarios": nombres})


@app.route('/api/proyeccion/rutas')
@limiter.limit("60 per minute")
def api_proyeccion_rutas():
    if not session.get('map_access'):
        return jsonify({"error": "No autorizado"}), 401
    escenario = (request.args.get('escenario') or 'Base').strip() or 'Base'
    cfg = _proyeccion_config(escenario)
    rutas = ProyeccionRuta.query.filter_by(escenario=escenario).order_by(
        ProyeccionRuta.desde_anio, ProyeccionRuta.desde_mes, ProyeccionRuta.id).all()

    _all_airports = Airport.query.all()
    airports = {a.name: (a.lat, a.lon) for a in _all_airports}
    es_argentino = {a.name: a.is_argentina for a in _all_airports}

    out = []
    for r in rutas:
        d = _ruta_proyeccion_to_dict(r)
        if r.origin in airports and r.dest in airports:
            o_ar = es_argentino.get(r.origin, False)
            d_ar = es_argentino.get(r.dest, False)
            factor = FACTOR_UPLIFT_CABOTAJE if (o_ar and d_ar) else FACTOR_UPLIFT_INTERNACIONAL
            dist = avion_model.haversine(*airports[r.origin], *airports[r.dest])
            if r.avion_forzado and r.avion_forzado in avion_model.get_flota():
                key = avion_model._key(r.origin, r.dest)
                tons, _o = avion_model.consumo_toneladas(r.avion_forzado, dist, key)
                fac, _lf = avion_model.ajuste_por_ocupacion(r.avion_forzado, r.pax_por_vuelo)
                m3 = ((tons or 0.0) * 1000.0 * fac) / avion_model.JETA1_KG_PER_M3
                nombre = avion_model.get_flota()[r.avion_forzado]['nombre']
            else:
                info = avion_model.get_aircraft_info(r.origin, r.dest, r.tipo, airports,
                                                     pax_por_vuelo=r.pax_por_vuelo)
                m3 = info['consumo_total_m3'] if info else 0.0
                nombre = info['avion'] if info else None
            d['avion'] = nombre
            d['distancia_km'] = round(dist, 1)
            d['factor_uplift'] = factor
            d['m3_por_vuelo'] = round(m3, 3)
            d['m3_por_mes'] = round(m3 * (r.vuelos_mes or 0) * factor * r.signo, 1)
        else:
            d['avion'] = None
            d['m3_por_mes'] = 0.0
        out.append(d)

    return jsonify({
        "escenario": escenario,
        "rutas": out,
        "config": {
            'crecimiento_anual_cabotaje': cfg.crecimiento_anual_cabotaje,
            'crecimiento_anual_intl': cfg.crecimiento_anual_intl,
            'horizonte_meses': cfg.horizonte_meses,
            'anios_estacionalidad': cfg.anios_estacionalidad,
            'nota': cfg.nota,
        },
        "aeropuertos": sorted(airports.keys()),
        "argentinos": sorted(n for n, v in es_argentino.items() if v),
        "flota": avion_model.get_flota(),
    })


def _parse_ruta_proyeccion_payload(data):
    """Valida el payload de alta/edición. Devuelve (campos, error)."""
    origin = (data.get('origin') or '').strip()
    dest = (data.get('dest') or '').strip()
    if not origin or not dest:
        return None, "Falta el origen o el destino."
    if origin == dest:
        return None, "El origen y el destino no pueden ser el mismo."

    airport_names = {n for (n,) in db.session.query(Airport.name).all()}
    faltan = [x for x in (origin, dest) if x not in airport_names]
    if faltan:
        return None, f"Aeropuerto desconocido: {', '.join(faltan)}"

    ar = {n for (n,) in db.session.query(Airport.name).filter(Airport.is_argentina.is_(True))}
    if origin not in ar and dest not in ar:
        return None, "La ruta no toca ningún aeropuerto argentino, no aporta combustible acá."
    tipo = 'cabotaje' if (origin in ar and dest in ar) else 'internacional'

    def _int(name, default=None, lo=None, hi=None):
        v = data.get(name, default)
        if v in (None, ''):
            return default
        try:
            v = int(v)
        except (TypeError, ValueError):
            raise ValueError(f"'{name}' tiene que ser un número entero.")
        if lo is not None and v < lo:
            raise ValueError(f"'{name}' no puede ser menor que {lo}.")
        if hi is not None and v > hi:
            raise ValueError(f"'{name}' no puede ser mayor que {hi}.")
        return v

    try:
        vuelos_mes = _int('vuelos_mes', 0, lo=0, hi=100000)
        desde_anio = _int('desde_anio', lo=2000, hi=2100)
        desde_mes = _int('desde_mes', lo=1, hi=12)
        hasta_anio = _int('hasta_anio', None, lo=2000, hi=2100)
        hasta_mes = _int('hasta_mes', None, lo=1, hi=12)
        signo = _int('signo', 1)
    except ValueError as e:
        return None, str(e)

    if desde_anio is None or desde_mes is None:
        return None, "Falta el mes de inicio."
    if signo not in (1, -1):
        return None, "El signo tiene que ser +1 (alta) o -1 (baja)."
    if (hasta_anio is None) != (hasta_mes is None):
        return None, "El mes de fin necesita año y mes, o ninguno de los dos."
    if hasta_anio is not None and (hasta_anio, hasta_mes) < (desde_anio, desde_mes):
        return None, "El mes de fin no puede ser anterior al de inicio."

    try:
        ppv = data.get('pax_por_vuelo')
        ppv = float(ppv) if ppv not in (None, '') else None
    except (TypeError, ValueError):
        return None, "'pax_por_vuelo' tiene que ser un número."
    if ppv is not None and not (0 < ppv <= 900):
        return None, "'pax_por_vuelo' tiene que estar entre 1 y 900."

    avion = (data.get('avion_forzado') or '').strip() or None
    if avion and avion not in avion_model.get_flota():
        return None, f"El avión '{avion}' no está en el catálogo de flota."

    return {
        'escenario': (data.get('escenario') or 'Base').strip()[:60] or 'Base',
        'origin': origin, 'dest': dest, 'tipo': tipo,
        'vuelos_mes': vuelos_mes, 'pax_por_vuelo': ppv, 'avion_forzado': avion,
        'signo': signo, 'desde_anio': desde_anio, 'desde_mes': desde_mes,
        'hasta_anio': hasta_anio, 'hasta_mes': hasta_mes,
        'activo': bool(data.get('activo', True)),
        'nota': (data.get('nota') or '').strip()[:500] or None,
    }, None


@app.route('/api/proyeccion/rutas/guardar', methods=['POST'])
@limiter.limit("30 per minute")
def api_proyeccion_ruta_guardar():
    """Alta o edición de una ruta proyectada. Alcanza con tener acceso al mapa: es un
    escenario hipotético, no toca ningún dato real de ANAC ni de combustible."""
    if not session.get('map_access'):
        return jsonify({"error": "No autorizado"}), 401
    try:
        data = request.get_json(silent=True) or {}
        campos, error = _parse_ruta_proyeccion_payload(data)
        if error:
            return jsonify({"error": error}), 400

        rid = data.get('id')
        if rid:
            r = ProyeccionRuta.query.get(int(rid))
            if r is None:
                return jsonify({"error": "No existe esa ruta proyectada."}), 404
            for k, v in campos.items():
                setattr(r, k, v)
        else:
            r = ProyeccionRuta(**campos)
            r.creado_por = session.get('user_email') or session.get('user_name')
            db.session.add(r)
        db.session.commit()
        return jsonify({"ok": True, "ruta": _ruta_proyeccion_to_dict(r)})
    except Exception as e:
        db.session.rollback()
        app.logger.exception("Error en /api/proyeccion/rutas/guardar")
        return jsonify({"error": f"Error inesperado en el servidor: {e}"}), 500


@app.route('/api/proyeccion/rutas/<int:rid>/borrar', methods=['POST'])
@limiter.limit("30 per minute")
def api_proyeccion_ruta_borrar(rid):
    if not session.get('map_access'):
        return jsonify({"error": "No autorizado"}), 401
    try:
        r = ProyeccionRuta.query.get(rid)
        if r is None:
            return jsonify({"error": "No existe esa ruta proyectada."}), 404
        db.session.delete(r)
        db.session.commit()
        return jsonify({"ok": True})
    except Exception as e:
        db.session.rollback()
        app.logger.exception("Error en /api/proyeccion/rutas/borrar")
        return jsonify({"error": f"Error inesperado en el servidor: {e}"}), 500


def _exclusion_to_dict(ex):
    return {
        'id': ex.id, 'aeropuerto': ex.aeropuerto,
        'desde_anio': ex.desde_anio, 'desde_mes': ex.desde_mes,
        'hasta_anio': ex.hasta_anio, 'hasta_mes': ex.hasta_mes,
        'motivo': ex.motivo, 'activo': ex.activo, 'creado_por': ex.creado_por,
    }


def _parse_exclusion_payload(data):
    """Valida el payload de alta/edición de un evento excluido. Devuelve (campos, error)."""
    aeropuerto = (data.get('aeropuerto') or '').strip() or None
    if aeropuerto:
        ok = db.session.query(Airport.name).filter(
            Airport.name == aeropuerto, Airport.is_argentina.is_(True)).first()
        if not ok:
            return None, f"'{aeropuerto}' no es un aeropuerto argentino conocido."

    motivo = (data.get('motivo') or '').strip()[:200]
    if not motivo:
        return None, "Falta el motivo (ej. 'Cierre de pista por obras')."

    def _int(name, lo, hi):
        v = data.get(name)
        try:
            v = int(v)
        except (TypeError, ValueError):
            raise ValueError(f"'{name}' tiene que ser un número entero.")
        if not (lo <= v <= hi):
            raise ValueError(f"'{name}' tiene que estar entre {lo} y {hi}.")
        return v

    try:
        desde_anio = _int('desde_anio', 2000, 2100)
        desde_mes = _int('desde_mes', 1, 12)
        hasta_anio = _int('hasta_anio', 2000, 2100)
        hasta_mes = _int('hasta_mes', 1, 12)
    except ValueError as e:
        return None, str(e)

    if (hasta_anio, hasta_mes) < (desde_anio, desde_mes):
        return None, "El mes de fin no puede ser anterior al de inicio."

    return {
        'aeropuerto': aeropuerto, 'motivo': motivo,
        'desde_anio': desde_anio, 'desde_mes': desde_mes,
        'hasta_anio': hasta_anio, 'hasta_mes': hasta_mes,
        'activo': bool(data.get('activo', True)),
    }, None


@app.route('/api/proyeccion/exclusiones')
@limiter.limit("60 per minute")
def api_proyeccion_exclusiones():
    if not session.get('map_access'):
        return jsonify({"error": "No autorizado"}), 401
    exclusiones = ProyeccionExclusion.query.order_by(
        ProyeccionExclusion.desde_anio, ProyeccionExclusion.desde_mes).all()
    return jsonify({"exclusiones": [_exclusion_to_dict(ex) for ex in exclusiones]})


@app.route('/api/proyeccion/exclusiones/guardar', methods=['POST'])
@limiter.limit("30 per minute")
def api_proyeccion_exclusion_guardar():
    """Alta o edición de una ventana de fechas a excluir del cálculo de estacionalidad y del
    nivel base. Igual que las rutas del escenario, alcanza con tener acceso al mapa: es una
    corrección sobre cómo se LEE el historial real, no un dato nuevo, y cualquiera que esté
    armando una proyección necesita poder ajustarla."""
    if not session.get('map_access'):
        return jsonify({"error": "No autorizado"}), 401
    try:
        data = request.get_json(silent=True) or {}
        campos, error = _parse_exclusion_payload(data)
        if error:
            return jsonify({"error": error}), 400

        eid = data.get('id')
        if eid:
            ex = ProyeccionExclusion.query.get(int(eid))
            if ex is None:
                return jsonify({"error": "No existe ese evento excluido."}), 404
            for k, v in campos.items():
                setattr(ex, k, v)
        else:
            ex = ProyeccionExclusion(**campos)
            ex.creado_por = session.get('user_email') or session.get('user_name')
            db.session.add(ex)
        db.session.commit()
        return jsonify({"ok": True, "exclusion": _exclusion_to_dict(ex)})
    except Exception as e:
        db.session.rollback()
        app.logger.exception("Error en /api/proyeccion/exclusiones/guardar")
        return jsonify({"error": f"Error inesperado en el servidor: {e}"}), 500


@app.route('/api/proyeccion/exclusiones/<int:eid>/borrar', methods=['POST'])
@limiter.limit("30 per minute")
def api_proyeccion_exclusion_borrar(eid):
    if not session.get('map_access'):
        return jsonify({"error": "No autorizado"}), 401
    try:
        ex = ProyeccionExclusion.query.get(eid)
        if ex is None:
            return jsonify({"error": "No existe ese evento excluido."}), 404
        db.session.delete(ex)
        db.session.commit()
        return jsonify({"ok": True})
    except Exception as e:
        db.session.rollback()
        app.logger.exception("Error en /api/proyeccion/exclusiones/borrar")
        return jsonify({"error": f"Error inesperado en el servidor: {e}"}), 500


_MESES_NOMBRE_A_NUM = {
    'ENERO': 1, 'FEBRERO': 2, 'MARZO': 3, 'ABRIL': 4, 'MAYO': 5, 'JUNIO': 6,
    'JULIO': 7, 'AGOSTO': 8, 'SEPTIEMBRE': 9, 'SETIEMBRE': 9, 'OCTUBRE': 10,
    'NOVIEMBRE': 11, 'DICIEMBRE': 12,
}


def parse_mercado_excel(file_obj):
    """Lee la planilla de mercado real. Columnas esperadas, en este orden:
    A = MES (nombre en castellano), B = AÑO, C = TOTAL, D = YPF, E = AXION, F = SHELL.

    Las filas de meses que todavía no tienen datos (típico: el año en curso, con los meses
    futuros en blanco) se saltean en silencio en vez de romper: la planilla suele tener los
    12 meses del año precargados y se van completando mes a mes.

    Devuelve (registros, warnings)."""
    wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
    ws = wb.active

    registros, warnings = [], []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if not row or len(row) < 3:
            continue
        vals = (list(row) + [None] * 6)[:6]
        mes_raw, anio_raw, total, ypf, axion, shell = vals

        mes_txt = _normalize_airport_name(str(mes_raw)) if mes_raw else ''
        mes = _MESES_NOMBRE_A_NUM.get(mes_txt)
        if not mes:
            continue  # encabezado o fila que no es de datos

        try:
            anio = int(anio_raw)
        except (TypeError, ValueError):
            warnings.append(f"Fila {i}: mes '{mes_raw}' sin año válido, se saltea.")
            continue

        if total in (None, ''):
            continue  # mes precargado sin datos todavía: no es un error

        def _num(v):
            try:
                return float(v) if v not in (None, '') else None
            except (TypeError, ValueError):
                return None

        total_v = _num(total)
        if total_v is None or total_v <= 0:
            warnings.append(f"Fila {i}: TOTAL no numérico o <= 0 para {mes_raw} {anio}, se saltea.")
            continue

        registros.append({
            'anio': anio, 'mes': mes, 'total_m3': total_v,
            'ypf_m3': _num(ypf), 'axion_m3': _num(axion), 'shell_m3': _num(shell),
        })
    return registros, warnings


@app.route('/api/cobertura_anac')
@limiter.limit("20 per minute")
def api_cobertura_anac():
    """Qué meses de tráfico de ANAC hay cargados, año por año, y de dónde salen.

    Sirve para responder "¿por qué la estacionalidad saltea 2023?" sin tener que adivinar:
    un año se saltea solo si tiene menos de 12 meses, y acá se ve exactamente cuáles faltan
    y si el año viene del histórico embebido o de un Excel subido al panel de admin."""
    if not session.get('map_access'):
        return jsonify({"error": "No autorizado"}), 401
    try:
        hist_routes, _ = get_historical_rows()

        meses_hist = {}   # anio -> set(mes)
        for r in hist_routes:
            meses_hist.setdefault(r['year'], set()).add(r['month'])

        meses_db = {}     # anio -> set(mes)
        for year, month in db.session.query(RouteMonthly.year, RouteMonthly.month).distinct():
            meses_db.setdefault(year, set()).add(month)

        anios = sorted(set(meses_hist) | set(meses_db), key=lambda a: int(a), reverse=True)
        out = []
        for anio in anios:
            h = meses_hist.get(anio, set())
            d = meses_db.get(anio, set())
            todos = h | d
            faltantes = [m for m in MONTH_ORDER if m not in todos]
            if h and d:
                origen = 'histórico + Excel subido'
            elif h:
                origen = 'histórico embebido'
            else:
                origen = 'Excel subido al admin'
            out.append({
                'anio': anio,
                'meses': len(todos),
                'completo': len(todos) >= 12,
                'faltantes': faltantes,
                'origen': origen,
            })
        return jsonify({
            "cobertura": out, "month_order": MONTH_ORDER,
            "descartadas_por_calidad": _HISTORICAL_DESCARTADAS,
        })
    except Exception as e:
        app.logger.exception("Error en /api/cobertura_anac")
        return jsonify({"error": str(e), "type": type(e).__name__}), 500


def parse_tipo_cambio_excel(file_obj):
    """Lee una planilla de tipo de cambio, en cualquiera de dos formatos:

    1) Propio: columna A = MES (nombre en castellano), B = AÑO, C = VALOR.
    2) Nativo del BCRA (tal cual se descarga de bcra.gob.ar, "Índices de Tipo de Cambio
       Multilateral" → ITCRM, sin tocar nada): columna A = fecha (fin de mes), columna B =
       valor del índice. Las columnas siguientes (ITCRB por país) se ignoran.

    Se detecta el formato mirando el tipo de dato de la columna A de cada fila: si es una
    fecha, formato BCRA; si es un nombre de mes en castellano, formato propio. Así el mismo
    botón sirve para el Excel que uno arma a mano y para el que se baja directo del banco
    central, sin tener que reformatear nada.

    Si el libro tiene varias hojas (como el del BCRA: la serie diaria y el promedio
    mensual), se prioriza una hoja cuyo nombre sugiera "mensual" — la diaria tiene ~30 veces
    más filas y no es la granularidad que usa el resto de la app.

    Mismo criterio que parse_mercado_excel: las filas de meses todavía sin datos, o las de
    encabezado/notas al pie, se saltean en silencio."""
    wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
    ws = wb.active
    for name in wb.sheetnames:
        low = name.lower()
        if 'mens' in low or 'mont' in low:
            ws = wb[name]
            break

    registros, warnings = [], []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if not row or len(row) < 2:
            continue

        primera = row[0]

        if hasattr(primera, 'year') and hasattr(primera, 'month'):
            # Formato BCRA: A = fecha, B = valor del índice.
            anio, mes = primera.year, primera.month
            valor_raw = row[1] if len(row) > 1 else None
        else:
            # Formato propio: A = mes en castellano, B = año, C = valor.
            if len(row) < 3:
                continue
            mes_txt = _normalize_airport_name(str(primera)) if primera else ''
            mes = _MESES_NOMBRE_A_NUM.get(mes_txt)
            if not mes:
                continue  # encabezado, nota al pie, o fila que no es de datos
            try:
                anio = int(row[1])
            except (TypeError, ValueError):
                warnings.append(f"Fila {i}: mes '{primera}' sin año válido, se saltea.")
                continue
            valor_raw = row[2]

        if valor_raw in (None, ''):
            continue  # mes precargado sin datos todavía, o fila de nota al pie: no es un error

        try:
            valor = float(valor_raw)
        except (TypeError, ValueError):
            warnings.append(f"Fila {i}: VALOR no numérico para {anio}-{mes:02d}, se saltea.")
            continue
        if valor <= 0:
            warnings.append(f"Fila {i}: VALOR <= 0 para {anio}-{mes:02d}, se saltea.")
            continue

        registros.append({'anio': anio, 'mes': mes, 'valor': valor})

    # El BCRA a veces reporta el mismo (año, mes) dos veces si la serie mensual arrastra un
    # dato "provisorio" y después lo revisa en la fila siguiente — nos quedamos con el
    # último valor de cada (año, mes) en vez de que bulk-upsert dependa del orden de un dict.
    por_periodo = {}
    for r in registros:
        por_periodo[(r['anio'], r['mes'])] = r
    return list(por_periodo.values()), warnings


@app.route('/api/proyeccion/fx')
@limiter.limit("60 per minute")
def api_proyeccion_fx():
    if not session.get('map_access'):
        return jsonify({"error": "No autorizado"}), 401
    filas = TipoCambioMensual.query.order_by(TipoCambioMensual.anio, TipoCambioMensual.mes).all()
    return jsonify({"fx": [{'anio': f.anio, 'mes': f.mes, 'valor': f.valor} for f in filas]})


@app.route('/api/proyeccion/fx/upload', methods=['POST'])
@limiter.limit("10 per minute")
def api_proyeccion_fx_upload():
    """Carga la planilla de tipo de cambio. Igual que el mercado real, va acá y no en
    /admin porque alimenta directamente la estimación de elasticidad de la proyección."""
    if not session.get('map_access'):
        return jsonify({"error": "No autorizado"}), 401
    try:
        f = request.files.get('file')
        if not f or not f.filename.lower().endswith(('.xlsx', '.xlsm')):
            return jsonify({"error": "Subí un archivo Excel (.xlsx) válido"}), 400

        try:
            registros, warnings = parse_tipo_cambio_excel(f)
        except Exception as e:
            return jsonify({"error": f"No se pudo leer el Excel: {e}"}), 400

        if not registros:
            return jsonify({"error": "No se encontró ninguna fila con MES, AÑO y VALOR. "
                                     "Se esperan las columnas A=MES, B=AÑO, C=VALOR."}), 400

        existentes = {(f.anio, f.mes): f for f in TipoCambioMensual.query.all()}
        insertados, actualizados = 0, 0
        for r in registros:
            f = existentes.get((r['anio'], r['mes']))
            if f:
                f.valor = r['valor']
                actualizados += 1
            else:
                db.session.add(TipoCambioMensual(**r))
                insertados += 1
        db.session.commit()

        return jsonify({"ok": True, "insertados": insertados, "actualizados": actualizados,
                        "warnings": warnings})
    except Exception as e:
        db.session.rollback()
        app.logger.exception("Error en /api/proyeccion/fx/upload")
        return jsonify({"error": f"Error inesperado en el servidor: {e}"}), 500


def parse_indec_csv(file_obj, categoria):
    """Lee un CSV publicado por INDEC, en cualquiera de dos formatos:

    1) ANCHO (índice de salarios): columna 'periodo' (fecha d/m/aaaa) + N columnas de
       valores, una por sub-serie, con 'NA' para los meses sin dato en una sub-serie que
       empezó más tarde que las demás.
    2) LARGO (IPC por división, serie_ipc_divisiones.csv): columnas Codigo;Descripcion;
       Clasificador;Periodo;Indice_IPC;v_m_IPC;v_i_a_IPC;Region, con el período como
       AAAAMM y una fila por (división, período, región) — la misma división aparece
       repetida una vez por cada una de las 7 regiones que publica INDEC.

    Se detecta cuál es mirando el encabezado: si trae 'Periodo', 'Region' y alguna
    columna con 'indice' en el nombre, es el formato largo; si no, es el ancho. En ambos
    casos separador ';', coma como separador decimal (configuración regional argentina).

    Cada sub-serie (columna en el formato ancho, división en el formato largo) se guarda
    bajo `categoria`, con su propio `nombre`, para poder combinarlas después (ej. salario
    real = salario nominal ÷ IPC, período a período)."""
    texto = file_obj.read()
    if isinstance(texto, bytes):
        # El índice de salarios que se venía cargando siempre es UTF-8, pero el IPC tal
        # cual lo publica INDEC (serie_ipc_divisiones.csv) viene en latin-1 (ISO-8859-1):
        # con errors='replace' a secas, cada tilde o ñ se pierde en silencio y se convierte
        # en '�' — el archivo se procesa "bien" (0 warnings) pero con los nombres de
        # división corrompidos. Se prueba UTF-8 estricto primero (falla ruidoso si no lo
        # es) y se cae a latin-1, que acepta cualquier secuencia de bytes sin perder nada.
        try:
            texto = texto.decode('utf-8-sig')
        except UnicodeDecodeError:
            texto = texto.decode('latin-1')
    lector = csv.reader(io.StringIO(texto), delimiter=';')

    filas = list(lector)
    if not filas:
        return [], ["El archivo está vacío."]
    encabezado = [h.strip() for h in filas[0]]
    if len(encabezado) < 2:
        return [], ["Se esperaba 'periodo' + al menos una columna de valores, separados por ';'."]

    enc_lower = [h.lower() for h in encabezado]
    if 'periodo' in enc_lower and 'region' in enc_lower and any('indice' in h for h in enc_lower):
        return _parse_indec_formato_largo(filas, encabezado, categoria)

    columnas = encabezado[1:]
    # prefijo común (ej. "IS_") se recorta del nombre guardado, para que quede legible
    prefijo = os.path.commonprefix(columnas) if len(columnas) > 1 else ''
    if not prefijo.endswith('_'):
        prefijo = ''
    nombres = [c[len(prefijo):] if prefijo and c.startswith(prefijo) else c for c in columnas]

    registros, warnings = [], []
    for i, fila in enumerate(filas[1:], start=2):
        if not fila or not fila[0].strip():
            continue
        periodo_raw = fila[0].strip()
        try:
            dia, mes, anio = (int(x) for x in periodo_raw.split('/'))
        except (ValueError, AttributeError):
            warnings.append(f"Fila {i}: no se pudo leer la fecha '{periodo_raw}', se saltea.")
            continue
        if not (1 <= mes <= 12):
            warnings.append(f"Fila {i}: mes fuera de rango en '{periodo_raw}', se saltea.")
            continue

        for j, nombre in enumerate(nombres):
            crudo = fila[j + 1].strip() if j + 1 < len(fila) else ''
            if not crudo or crudo.upper() == 'NA':
                continue  # sub-serie que todavía no arrancó ese mes: no es un error
            try:
                valor = float(crudo.replace(',', '.'))
            except ValueError:
                warnings.append(f"Fila {i}, columna '{nombre}': '{crudo}' no es numérico, se saltea.")
                continue
            registros.append({'categoria': categoria, 'nombre': nombre,
                              'anio': anio, 'mes': mes, 'valor': valor})

    # último valor gana si el archivo repite un (nombre, año, mes) — igual criterio que FX
    por_clave = {}
    for r in registros:
        por_clave[(r['nombre'], r['anio'], r['mes'])] = r
    return list(por_clave.values()), warnings


def _parse_indec_formato_largo(filas, encabezado, categoria):
    """CSV 'largo' estilo IPC de INDEC: Codigo;Descripcion;Clasificador;Periodo;
    Indice_IPC;v_m_IPC;v_i_a_IPC;Region.

    Se filtra a Region='Nacional': el archivo trae 7 aperturas regionales (GBA, Pampeana,
    Cuyo, Noreste, Noroeste, Patagonia) más el total nacional, y para deflactar una serie
    nacional de tráfico aéreo la apertura regional no aporta nada — sin el filtro, cada
    división quedaría repetida 7 veces con valores distintos bajo el mismo `nombre`, y
    (categoria, nombre, año, mes) dejaría de ser una clave única.

    Cada división de la tabla COICOP (Codigo/Descripcion: '0' Nivel general, '01'..'12'
    las 12 divisiones, más los agregados 'Estacional', 'Núcleo', 'Regulados', 'B', 'S') se
    guarda como una sub-serie separada, con `nombre` en Title Case en vez del código
    COICOP crudo."""
    idx = {h.strip().lower(): i for i, h in enumerate(encabezado)}
    requeridas = ('codigo', 'descripcion', 'periodo', 'indice_ipc', 'region')
    faltan = [c for c in requeridas if c not in idx]
    if faltan:
        return [], [f"Formato largo detectado pero faltan columnas: {', '.join(faltan)}."]

    registros, warnings = [], []
    for i, fila in enumerate(filas[1:], start=2):
        if not fila or len(fila) <= max(idx.values()):
            continue
        if fila[idx['region']].strip() != 'Nacional':
            continue  # aperturas regionales: no interesan para un deflactor nacional

        periodo_raw = fila[idx['periodo']].strip()
        if len(periodo_raw) != 6 or not periodo_raw.isdigit():
            warnings.append(f"Fila {i}: período '{periodo_raw}' no tiene formato AAAAMM, se saltea.")
            continue
        anio, mes = int(periodo_raw[:4]), int(periodo_raw[4:])
        if not (1 <= mes <= 12):
            warnings.append(f"Fila {i}: mes fuera de rango en '{periodo_raw}', se saltea.")
            continue

        crudo = fila[idx['indice_ipc']].strip()
        if not crudo or crudo.upper() == 'NA':
            continue  # división sin serie ese mes (agregados como 'Núcleo' arrancan después)
        try:
            valor = float(crudo.replace(',', '.'))
        except ValueError:
            warnings.append(f"Fila {i}: Indice_IPC '{crudo}' no es numérico, se saltea.")
            continue

        descripcion = fila[idx['descripcion']].strip()
        codigo = fila[idx['codigo']].strip()
        # Las 13 divisiones COICOP (código '0', '01'..'12') traen Descripcion; los 5
        # agregados especiales ('Estacional', 'Núcleo', 'Regulados', 'B', 'S') no — pero
        # su propio código ya es una etiqueta legible, así que se usa tal cual en vez de
        # anteponerle "Código" (que sólo tendría sentido para un código puramente numérico
        # sin ninguna etiqueta, caso que no ocurre hoy pero se cubre igual).
        if descripcion:
            nombre = descripcion.title()
        elif codigo and not codigo.isdigit():
            nombre = codigo
        else:
            nombre = f"Código {codigo}"
        registros.append({'categoria': categoria, 'nombre': nombre,
                          'anio': anio, 'mes': mes, 'valor': valor})

    por_clave = {}
    for r in registros:
        por_clave[(r['nombre'], r['anio'], r['mes'])] = r
    return list(por_clave.values()), warnings


# Mapeo manual de abreviatura de mes en inglés -> número, en vez de confiar en
# datetime.strptime('%b', ...) que depende del locale del sistema operativo (en un locale
# no inglés, "Aug" no matchea). Así el parseo es determinístico sin importar dónde corra.
_MES_ABR_EN = {'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'May': 5, 'Jun': 6,
               'Jul': 7, 'Aug': 8, 'Sep': 9, 'Oct': 10, 'Nov': 11, 'Dec': 12}


def parse_dolar_blue_csv(file_obj):
    """CSV diario tal cual lo exportan varios trackers del dólar blue: encabezado
    'category,valor', fecha en inglés y en su propia columna ('Mon Aug 08 2016'),
    separador coma, PUNTO decimal — el único CSV de todo el proyecto que no viene en
    formato argentino (los de INDEC usan ';' y coma decimal).

    Se promedian los valores diarios de cada mes calendario para bajar la serie a
    granularidad mensual, que es con la que trabaja el resto de la app. Se guarda como
    IndiceEconomico(categoria='dolar_blue', nombre='nominal'): mismo mecanismo genérico que
    salarios e IPC, así que /api/proyeccion/indices?categoria=dolar_blue ya sirve para
    leerlo de vuelta sin agregar un endpoint GET nuevo."""
    texto = file_obj.read()
    if isinstance(texto, bytes):
        try:
            texto = texto.decode('utf-8-sig')
        except UnicodeDecodeError:
            texto = texto.decode('latin-1')
    lector = csv.reader(io.StringIO(texto))
    filas = list(lector)
    if not filas:
        return [], ["El archivo está vacío."]
    encabezado = [h.strip().lower() for h in filas[0]]
    if len(encabezado) < 2:
        return [], ["Se esperaban 2 columnas: fecha y valor, separadas por coma."]

    por_mes = {}  # (anio, mes) -> lista de valores diarios, para promediar
    warnings = []
    for i, fila in enumerate(filas[1:], start=2):
        if not fila or len(fila) < 2 or not fila[0].strip():
            continue
        fecha_raw = fila[0].strip()
        valor_raw = fila[1].strip()
        partes = fecha_raw.split()
        if len(partes) != 4 or partes[1] not in _MES_ABR_EN:
            warnings.append(f"Fila {i}: no se pudo leer la fecha '{fecha_raw}' (se espera "
                            f"'Dow Mon DD AAAA', ej. 'Mon Aug 08 2016'), se saltea.")
            continue
        mes = _MES_ABR_EN[partes[1]]
        try:
            anio = int(partes[3])
        except ValueError:
            warnings.append(f"Fila {i}: año inválido en '{fecha_raw}', se saltea.")
            continue
        try:
            valor = float(valor_raw)
        except ValueError:
            warnings.append(f"Fila {i}: valor '{valor_raw}' no es numérico, se saltea.")
            continue
        por_mes.setdefault((anio, mes), []).append(valor)

    registros = [
        {'categoria': 'dolar_blue', 'nombre': 'nominal', 'anio': anio, 'mes': mes,
         'valor': sum(vals) / len(vals)}
        for (anio, mes), vals in por_mes.items()
    ]
    return registros, warnings


@app.route('/api/proyeccion/indices')
@limiter.limit("60 per minute")
def api_proyeccion_indices():
    if not session.get('map_access'):
        return jsonify({"error": "No autorizado"}), 401
    categoria = (request.args.get('categoria') or '').strip() or None
    q = IndiceEconomico.query
    if categoria:
        q = q.filter_by(categoria=categoria)
    filas = q.order_by(IndiceEconomico.categoria, IndiceEconomico.nombre,
                       IndiceEconomico.anio, IndiceEconomico.mes).all()
    return jsonify({
        "indices": [{'categoria': f.categoria, 'nombre': f.nombre, 'anio': f.anio,
                     'mes': f.mes, 'valor': f.valor} for f in filas],
        "series_disponibles": sorted({(f.categoria, f.nombre) for f in filas}),
    })


@app.route('/api/proyeccion/indices/upload', methods=['POST'])
@limiter.limit("10 per minute")
def api_proyeccion_indices_upload():
    """Carga un CSV estilo INDEC (salarios, o cualquier otro con el mismo formato:
    'periodo;col1;col2;...'). La categoría viene por parámetro para poder distinguir
    salarios de otras series que se carguen más adelante (ej. IPC)."""
    if not session.get('map_access'):
        return jsonify({"error": "No autorizado"}), 401
    try:
        categoria = (request.form.get('categoria') or '').strip()
        if not categoria:
            return jsonify({"error": "Falta la categoría (ej. 'salario')."}), 400

        f = request.files.get('file')
        if not f or not f.filename.lower().endswith('.csv'):
            return jsonify({"error": "Subí un archivo CSV (.csv) válido"}), 400

        try:
            if categoria == 'dolar_blue':
                registros, warnings = parse_dolar_blue_csv(f)
            else:
                registros, warnings = parse_indec_csv(f, categoria)
        except Exception as e:
            return jsonify({"error": f"No se pudo leer el CSV: {e}"}), 400

        if not registros:
            if categoria == 'dolar_blue':
                return jsonify({"error": "No se encontró ninguna fila válida. Se espera "
                                         "'category,valor' con fecha en inglés (ej. 'Mon Aug "
                                         "08 2016') y punto decimal."}), 400
            return jsonify({"error": "No se encontró ninguna fila válida. Se espera "
                                     "'periodo;columna1;columna2;...' separado por ';', con "
                                     "fechas d/m/aaaa y coma como separador decimal."}), 400

        existentes = {(f.categoria, f.nombre, f.anio, f.mes): f
                      for f in IndiceEconomico.query.filter_by(categoria=categoria).all()}
        insertados, actualizados = 0, 0
        for r in registros:
            clave = (r['categoria'], r['nombre'], r['anio'], r['mes'])
            f_existente = existentes.get(clave)
            if f_existente:
                f_existente.valor = r['valor']
                actualizados += 1
            else:
                db.session.add(IndiceEconomico(**r))
                insertados += 1
        db.session.commit()

        series = sorted({r['nombre'] for r in registros})
        return jsonify({"ok": True, "insertados": insertados, "actualizados": actualizados,
                        "series": series, "warnings": warnings})
    except Exception as e:
        db.session.rollback()
        app.logger.exception("Error en /api/proyeccion/indices/upload")
        return jsonify({"error": f"Error inesperado en el servidor: {e}"}), 500


@app.route('/api/proyeccion/mercado')
@limiter.limit("60 per minute")
def api_proyeccion_mercado():
    if not session.get('map_access'):
        return jsonify({"error": "No autorizado"}), 401
    filas = MercadoMensual.query.order_by(MercadoMensual.anio, MercadoMensual.mes).all()
    return jsonify({"mercado": [{
        'anio': m.anio, 'mes': m.mes, 'total_m3': m.total_m3,
        'ypf_m3': m.ypf_m3, 'axion_m3': m.axion_m3, 'shell_m3': m.shell_m3,
        'otros_m3': round(m.otros_m3, 1),
        'share_ypf': round(m.ypf_m3 / m.total_m3, 4) if (m.ypf_m3 and m.total_m3) else None,
    } for m in filas]})


@app.route('/api/proyeccion/mercado/upload', methods=['POST'])
@limiter.limit("10 per minute")
def api_proyeccion_mercado_upload():
    """Carga la planilla de mercado real por empresa. Va acá y no en /admin porque es el dato
    que cierra el circuito de la proyección: sin él, el desvío del modelo no se puede medir."""
    if not session.get('map_access'):
        return jsonify({"error": "No autorizado"}), 401
    try:
        f = request.files.get('file')
        if not f or not f.filename.lower().endswith(('.xlsx', '.xlsm')):
            return jsonify({"error": "Subí un archivo Excel (.xlsx) válido"}), 400

        try:
            registros, warnings = parse_mercado_excel(f)
        except Exception as e:
            return jsonify({"error": f"No se pudo leer el Excel: {e}"}), 400

        if not registros:
            return jsonify({"error": "No se encontró ninguna fila con MES, AÑO y TOTAL. "
                                     "Se esperan las columnas A=MES, B=AÑO, C=TOTAL, "
                                     "D=YPF, E=AXION, F=SHELL."}), 400

        existentes = {(m.anio, m.mes): m for m in MercadoMensual.query.all()}
        insertados, actualizados = 0, 0
        for r in registros:
            m = existentes.get((r['anio'], r['mes']))
            if m:
                m.total_m3 = r['total_m3']
                m.ypf_m3 = r['ypf_m3']
                m.axion_m3 = r['axion_m3']
                m.shell_m3 = r['shell_m3']
                actualizados += 1
            else:
                db.session.add(MercadoMensual(**r))
                insertados += 1
        db.session.commit()

        return jsonify({"ok": True, "insertados": insertados, "actualizados": actualizados,
                        "warnings": warnings})
    except Exception as e:
        db.session.rollback()
        app.logger.exception("Error en /api/proyeccion/mercado/upload")
        return jsonify({"error": f"Error inesperado en el servidor: {e}"}), 500


@app.route('/api/proyeccion/config', methods=['POST'])
@limiter.limit("30 per minute")
def api_proyeccion_config():
    if not session.get('map_access'):
        return jsonify({"error": "No autorizado"}), 401
    try:
        data = request.get_json(silent=True) or {}
        escenario = (data.get('escenario') or 'Base').strip()[:60] or 'Base'
        cfg = _proyeccion_config(escenario)

        def _num(name, actual, lo, hi):
            v = data.get(name)
            if v in (None, ''):
                return actual
            try:
                v = float(v)
            except (TypeError, ValueError):
                raise ValueError(f"'{name}' tiene que ser un número.")
            if not (lo <= v <= hi):
                raise ValueError(f"'{name}' tiene que estar entre {lo} y {hi}.")
            return v

        cfg.crecimiento_anual_cabotaje = _num('crecimiento_anual_cabotaje',
                                              cfg.crecimiento_anual_cabotaje, -0.5, 0.5)
        cfg.crecimiento_anual_intl = _num('crecimiento_anual_intl',
                                          cfg.crecimiento_anual_intl, -0.5, 0.5)
        cfg.horizonte_meses = int(_num('horizonte_meses', cfg.horizonte_meses, 1, 120))
        cfg.anios_estacionalidad = int(_num('anios_estacionalidad',
                                            cfg.anios_estacionalidad, 1, 10))
        if 'nota' in data:
            cfg.nota = (data.get('nota') or '').strip()[:500] or None
        db.session.commit()
        return jsonify({"ok": True})
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        db.session.rollback()
        app.logger.exception("Error en /api/proyeccion/config")
        return jsonify({"error": f"Error inesperado en el servidor: {e}"}), 500


if __name__ == '__main__':
    with app.app_context():
        ensure_tables()
    app.run(debug=True, port=5000)
