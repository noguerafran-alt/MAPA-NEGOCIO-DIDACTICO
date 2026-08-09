"""
Parser for ANAC "series históricas" Excel workbooks (sheet 'OUT').
Locates tables by their title text (not fixed row numbers) so it keeps
working even if ANAC inserts/removes tables in future releases.

Uses openpyxl directly in read-only/streaming mode (no pandas DataFrame)
to keep peak memory low — important on 512MB free hosting tiers.
"""
import re
import math
import openpyxl

MONTH_ORDER = ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]

TITLE_ROUTE_PAX_CABOTAJE = "Pasajeros Comerciales Cabotaje [000] (Pax)"
TITLE_ROUTE_PAX_INTL = "Pasajeros Comerciales Internacionales [000] (Pax)"
TITLE_ROUTE_VUELOS_CABOTAJE = "Vuelos Comerciales Cabotaje [#] (Sin cargueros exclusivos, ferrys, ni vuelos que regresan)"
TITLE_ROUTE_VUELOS_INTL = "Vuelos Comerciales Internacionales [#] (Sin cargueros exclusivos, ferrys, ni vuelos que regresan)"
TITLE_AIRPORT_PAX_TOTAL = "Pasajeros Totales [000]"

TABLA_RE = re.compile(r"^TABLA\s+\d+$")

REQUIRED_TITLES = {
    'cabotaje_pax': TITLE_ROUTE_PAX_CABOTAJE,
    'intl_pax': TITLE_ROUTE_PAX_INTL,
    'cabotaje_vuelos': TITLE_ROUTE_VUELOS_CABOTAJE,
    'intl_vuelos': TITLE_ROUTE_VUELOS_INTL,
    'airport_pax': TITLE_AIRPORT_PAX_TOTAL,
}


def safe_float(v):
    if v is None:
        return None
    try:
        f = float(v)
    except (TypeError, ValueError):
        return None
    if math.isnan(f):
        return None
    return f


def parse_workbook(file_path_or_buffer):
    """
    Parse one ANAC 'series históricas' xlsx file using streaming reads only.
    Returns dict with keys: cabotaje_pax, intl_pax, cabotaje_vuelos, intl_vuelos, airport_pax
    Each maps route/airport name -> {year-month: raw_value}
    """
    wb = openpyxl.load_workbook(file_path_or_buffer, read_only=True, data_only=True)
    ws = wb['OUT']

    # --- Pass 1: scan column A only to find "TABLA N" markers and their titles ---
    col_a = []
    for row in ws.iter_rows(min_col=1, max_col=1, values_only=True):
        col_a.append(row[0])

    table_start_rows = {}  # title -> 0-indexed row number (row of "TABLA N")
    for i in range(len(col_a) - 1):
        v = col_a[i]
        if isinstance(v, str) and TABLA_RE.match(v.strip()):
            title = col_a[i + 1]
            if isinstance(title, str):
                table_start_rows[title.strip()] = i

    missing = [key for key, title in REQUIRED_TITLES.items() if title not in table_start_rows]
    if missing:
        wb.close()
        raise ValueError(
            f"No se encontraron estas tablas en el archivo: {missing}. "
            f"Títulos detectados en el archivo: {sorted(table_start_rows.keys())}"
        )

    def find_data_end(start_row):
        r = start_row + 2
        while r < len(col_a) and col_a[r] is not None:
            r += 1
        return r  # exclusive

    table_ranges = {}  # key -> (start_row, data_start, data_end) all 0-indexed
    for key, title in REQUIRED_TITLES.items():
        start_row = table_start_rows[title]
        data_start = start_row + 2
        data_end = find_data_end(start_row)
        table_ranges[key] = (start_row, data_start, data_end)

    del col_a  # free before the heavier pass

    # --- Pass 2: for each needed table, stream just its row range and build the dict ---
    result = {key: {} for key in REQUIRED_TITLES}

    for key, (start_row, data_start, data_end) in table_ranges.items():
        # openpyxl rows are 1-indexed; +1 to convert from our 0-indexed row numbers.
        #
        # FIX (encontrado en agosto 2026): la fila "TABLA N" trae, en columna A, la
        # etiqueta "TABLA N", y en columna B nada (espaciador, usado solo por las
        # tablas con desglose Comercial/Privados más abajo en el archivo, donde
        # guarda "Regular/No Regular/Nacional/Extranjero"). El bloque de años/meses
        # y los datos arrancan siempre en la COLUMNA C — antes esto decía
        # `min_col=4` (columna D) y `row[3:3+n_cols]`, una columna de más a la
        # derecha. Como año y mes se leían con el mismo offset que los datos, el
        # resultado quedaba "autoconsistente" (cada mes seguía emparejado con su
        # valor correcto) y no tiraba ningún error — pero la PRIMERA columna del
        # archivo (el primer mes del primer año, ej. 2023-Ene) se descartaba en
        # silencio siempre, en las 5 tablas que lee este parser. Confirmado contra
        # un archivo real: sin el fix, 2023 quedaba con 11 meses (sin Ene) mientras
        # todos los demás años tenían los 12 completos.
        year_row = next(ws.iter_rows(min_row=start_row + 1, max_row=start_row + 1, min_col=3, values_only=True))
        month_row = next(ws.iter_rows(min_row=start_row + 2, max_row=start_row + 2, min_col=3, values_only=True))

        # Una columna sin año o sin mes NO corta la lectura: se marca como None para
        # conservar la alineación posicional con las celdas de datos, y se descarta al armar
        # el dict. Antes esto hacía `break`, así que una sola columna separadora en medio del
        # encabezado (algo que ANAC intercala entre bloques de años en algunos archivos)
        # descartaba en silencio TODOS los meses siguientes — un año entero podía
        # desaparecer sin ningún mensaje de error.
        cols = []
        for y, m in zip(year_row, month_row):
            if y is None or m is None:
                cols.append(None)
                continue
            try:
                cols.append(f"{int(y)}-{m}")
            except (TypeError, ValueError):
                cols.append(None)  # año no numérico: misma lógica que un hueco
        # Recortar las columnas vacías del final (después de la última con datos reales),
        # que sí son un corte legítimo y no hace falta leer.
        while cols and cols[-1] is None:
            cols.pop()
        n_cols = len(cols)

        if data_end > data_start:
            for row in ws.iter_rows(min_row=data_start + 1, max_row=data_end, min_col=1, max_col=2 + n_cols,
                                     values_only=True):
                name = row[0]
                if name is None:
                    continue
                vals = row[2:2 + n_cols]
                result[key][name] = {c: v for c, v in zip(cols, vals) if c is not None}

    wb.close()
    return result


def rows_to_monthly_records(table_dict):
    """Convert {name: {year-month: val}} -> list of (name, year, month, value)."""
    out = []
    for name, ymvals in table_dict.items():
        for ym, val in ymvals.items():
            year, month = ym.split('-', 1)
            v = safe_float(val)
            out.append((name.strip(), year, month, v))
    return out
